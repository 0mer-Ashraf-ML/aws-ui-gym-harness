"""Generate Excel reports summarising execution iterations across runners."""

from __future__ import annotations

import json
import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.database_utils import get_db_session
from app.models.iteration import Iteration

logger = logging.getLogger(__name__)


KNOWN_RUNNERS = {
    "anthropic": "Claude Sonnet 4",
    "openai": "OpenAI Computer Use Preview",
    "gemini": "Google Gemini Computer Use",
}

RUNNER_MODELS = {
    "anthropic": "Claude Sonnet 4",
    "openai": "OpenAI Computer Use Preview",
    "gemini": "Google Gemini Computer Use",
}

MODEL_ORDER = [
    ("anthropic", RUNNER_MODELS["anthropic"]),
    ("openai", RUNNER_MODELS["openai"]),
    ("gemini", RUNNER_MODELS["gemini"]),
]


@dataclass
class IterationRecord:
    task_id: str
    iteration: int
    runner: str
    status: str
    status_reason: Optional[str]
    completion_reason: Optional[str]
    duration_seconds: Optional[float]
    timelapse: Optional[str]
    file_timelapse_seconds: Optional[float]
    tool_calls_total: int
    tool_calls_by_tool: Dict[str, int] = field(default_factory=dict)
    unique_tools: List[str] = field(default_factory=list)
    prompt: Optional[str] = None
    prompt_id: Optional[str] = None
    model: Optional[str] = None
    run_id: Optional[str] = None
    start_timestamp: Optional[str] = None
    end_timestamp: Optional[str] = None
    iteration_directory: Optional[str] = None
    execution_uuid: Optional[str] = None
    iteration_uuid: Optional[str] = None
    verification_comments: Optional[str] = None
    last_model_response: Optional[str] = None  # Database field for model response
    eval_insights: Optional[str] = None  # Database field for evaluation insights
    total_steps: Optional[int] = None  # Database field for total steps
    extra: Dict[str, object] = field(default_factory=dict)

    @property
    def runner_label(self) -> str:
        return KNOWN_RUNNERS.get(self.runner, self.runner.title())


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_execution_report(
    execution_dir: Path | str,
    output_path: Optional[Path | str] = None,
    *,
    write_json: bool = True,
) -> Path:
    """Generate an Excel workbook summarising execution runs.

    Args:
        execution_dir: Directory containing per-task execution folders.
        output_path: Optional path for the generated workbook. Defaults to
            ``<execution_dir>/<execution_dir.name>_report.xlsx``.

    Returns:
        Path to the generated workbook.
    """

    execution_path = Path(execution_dir).expanduser().resolve()
    if not execution_path.exists():
        raise FileNotFoundError(f"Execution directory not found: {execution_path}")

    records = _collect_iteration_records(execution_path)
    if not records:
        raise ValueError(f"No iteration records found inside {execution_path}")

    summary_rows, task_rows = _build_summary(records)

    if output_path is None:
        output_path = execution_path / f"{execution_path.name}_report.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    _write_workbook(
        summary_rows=summary_rows,
        iterations=records,
        task_rows=task_rows,
        workbook_path=output_path,
    )

    if write_json:
        snapshot_path = output_path.with_suffix(".json")
        snapshot = _build_snapshot(summary_rows, records, task_rows)
        _write_json_snapshot(snapshot_path, snapshot)

    logger.info("📊 Execution report generated: %s", output_path)
    return output_path


def collect_execution_data(execution_dir: Path | str) -> List[IterationRecord]:
    """
    Collect iteration records from an execution directory.
    
    This is a public API for accessing the raw data used in report generation.
    Use this for JSON APIs to ensure DRY - single source of truth for data extraction.
    
    Args:
        execution_dir: Directory containing per-task execution folders.
        
    Returns:
        List of IterationRecord objects with all iteration data including:
        - task_id, iteration, runner, status
        - tool_calls_total, tool_calls_by_tool, unique_tools
        - duration, timelapse, timestamps
        - prompt, model, completion_reason, status_reason
    """
    execution_path = Path(execution_dir).expanduser().resolve()
    if not execution_path.exists():
        raise FileNotFoundError(f"Execution directory not found: {execution_path}")
    
    return _collect_iteration_records(execution_path)


def generate_combined_report(
    execution_dirs: Iterable[Path | str],
    output_path: Optional[Path | str] = None,
    *,
    write_json: bool = True,
) -> Path:
    """Generate a single workbook aggregating multiple execution directories."""

    dirs: List[Path] = []
    for directory in execution_dirs:
        path = Path(directory).expanduser().resolve()
        if path.exists() and path.is_dir():
            dirs.append(path)
        else:
            logger.warning("Skipping missing execution directory: %s", directory)

    if not dirs:
        raise ValueError("No valid execution directories provided")

    records: List[IterationRecord] = []
    for directory in dirs:
        try:
            records.extend(_collect_iteration_records(directory))
        except Exception as exc:  # pragma: no cover - defensive
            logger.warning("Failed to collect records from %s: %s", directory, exc)

    if not records:
        raise ValueError("No iteration data found across execution directories")

    summary_rows, task_rows = _build_summary(records)

    if output_path is None:
        output_path = dirs[0].parent / "combined_execution_report.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    _write_workbook(
        summary_rows=summary_rows,
        iterations=records,
        task_rows=task_rows,
        workbook_path=output_path,
    )

    if write_json:
        snapshot_path = output_path.with_suffix(".json")
        snapshot = _build_snapshot(summary_rows, records, task_rows)
        _write_json_snapshot(snapshot_path, snapshot)

    logger.info("📊 Combined execution report generated: %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Data collection helpers
# ---------------------------------------------------------------------------

def _get_total_steps_from_db(iteration_uuid: Optional[str]) -> Optional[int]:
    """Get total_steps from database if iteration_uuid is available."""
    if not iteration_uuid:
        return None
    try:
        import uuid as uuid_lib
        uuid_obj = uuid_lib.UUID(iteration_uuid)
        
        with get_db_session() as session:
            iteration = session.query(Iteration).filter(Iteration.uuid == uuid_obj).first()
            return iteration.total_steps if iteration and iteration.total_steps is not None else None
    except Exception:
        return None


def _collect_iteration_records(execution_dir: Path) -> List[IterationRecord]:
    records: List[IterationRecord] = []

    # Check if this is a batch directory (has different structure)
    if _is_batch_directory(execution_dir):
        return _collect_batch_iteration_records(execution_dir)

    # Standard execution directory structure
    for task_dir in sorted(p for p in execution_dir.iterdir() if p.is_dir()):
        task_id = task_dir.name
        for iteration_dir in sorted(p for p in task_dir.iterdir() if p.is_dir()):
            iteration_index = _extract_iteration_number(iteration_dir.name)
            for runner_dir in sorted(p for p in iteration_dir.iterdir() if p.is_dir()):
                runner_name = runner_dir.name.lower()
                record = _load_iteration_record(
                    task_id=task_id,
                    iteration_index=iteration_index,
                    runner_name=runner_name,
                    runner_path=runner_dir,
                )
                if record:
                    records.append(record)

    return records


def _is_batch_directory(execution_dir: Path) -> bool:
    """Check if this is a batch directory with the batch structure."""
    # Batch directories have this structure: batch_*/TASK_ID/iteration_N/
    # and iteration directories contain conversation_history, task_responses, etc.
    # instead of runner subdirectories
    
    for task_dir in execution_dir.iterdir():
        if not task_dir.is_dir():
            continue
        
        # Check if there are iteration directories
        iteration_dirs = [p for p in task_dir.iterdir() if p.is_dir() and p.name.startswith('iteration')]
        if not iteration_dirs:
            continue
            
        # Check the first iteration directory to see if it has batch structure
        first_iteration = iteration_dirs[0]
        batch_subdirs = {'conversation_history', 'task_responses', 'screenshots', 'logs'}
        actual_subdirs = {p.name for p in first_iteration.iterdir() if p.is_dir()}
        
        # If it has batch subdirectories and no runner subdirectories, it's a batch
        if batch_subdirs.intersection(actual_subdirs) and not any(runner in actual_subdirs for runner in ['anthropic', 'openai', 'gemini']):
            return True
    
    return False


def _collect_batch_iteration_records(execution_dir: Path) -> List[IterationRecord]:
    """Collect iteration records from batch directory structure."""
    records: List[IterationRecord] = []
    
    # Extract runner name from batch directory name
    batch_name = execution_dir.name
    runner_name = None
    if 'anthropic' in batch_name.lower():
        runner_name = 'anthropic'
    elif 'openai' in batch_name.lower():
        runner_name = 'openai'
    elif 'gemini' in batch_name.lower():
        runner_name = 'gemini'
    else:
        logger.warning("⚠️ Could not determine runner from batch name: %s", batch_name)
        return records
    
    for task_dir in sorted(p for p in execution_dir.iterdir() if p.is_dir()):
        task_id = task_dir.name
        for iteration_dir in sorted(p for p in task_dir.iterdir() if p.is_dir()):
            iteration_index = _extract_iteration_number(iteration_dir.name)
            
            # For batch directories, the iteration_dir IS the runner_path
            # since there are no runner subdirectories
            record = _load_iteration_record(
                task_id=task_id,
                iteration_index=iteration_index,
                runner_name=runner_name,
                runner_path=iteration_dir,  # iteration_dir is the runner_path for batches
            )
            if record:
                records.append(record)
    
    return records


def _load_batch_iteration_record(
    *,
    task_id: str,
    iteration_index: Optional[int],
    runner_name: str,
    runner_path: Path,
) -> Optional[IterationRecord]:
    """Load iteration record from batch directory structure (no iteration_meta.json)."""
    
    # Read verification_results.json for status and metadata
    verification_path = runner_path / "verification_results.json"
    verification = _safe_load_json(verification_path) or {}
    
    status_reason = verification.get("verification_summary")
    completion_reason: Optional[str] = None
    status = _infer_status(
        verification.get("verification_status"),
        extra=verification,
        status_reason=status_reason,
        completion_reason=completion_reason,
    )
    
    # Clean up status_reason if it contains HTML
    if status_reason and "<!doctype" in status_reason.lower():
        if ":" in status_reason:
            status_reason = status_reason.split(":", 1)[0].strip()
        else:
            status_reason = "Invalid verification response"
    
    run_id = verification.get("run_id")
    prompt_id = verification.get("prompt_id")
    
    # Extract model response from conversation history (PRIORITY)
    conversation_dir = runner_path / "conversation_history"
    if conversation_dir.exists():
        for conv_file in sorted(conversation_dir.glob("*_task_execution_conversation.json")):
            conv_data = _safe_load_json(conv_file) or {}
            conversation_flow = conv_data.get("conversation_flow", [])
            
            # Try conversation_flow first (newer format)
            if conversation_flow:
                for item in reversed(conversation_flow):
                    if item.get("role") == "assistant" and item.get("content"):
                        content = item["content"]
                        if content and content.strip():
                            completion_reason = content
                            break
                if completion_reason:
                    break
            
            # Fallback to messages array
            messages = conv_data if isinstance(conv_data, list) else conv_data.get("messages", [])
            for msg in reversed(messages):
                if msg.get("role") == "assistant" and msg.get("content"):
                    content = msg["content"]
                    if isinstance(content, str) and content.strip():
                        completion_reason = content
                        break
                    elif isinstance(content, list):
                        for block in content:
                            if isinstance(block, dict) and block.get("type") == "text":
                                text = block.get("text", "")
                                if text and text.strip():
                                    completion_reason = text
                                    break
            if completion_reason:
                break
    
    
    # Count tool calls from task_responses
    tool_calls_total = 0
    tool_calls_by_tool = {}
    
    task_responses_dir = runner_path / "task_responses"
    if task_responses_dir.exists():
        for response_file in sorted(task_responses_dir.glob("*.json")):
            response_data = _safe_load_json(response_file) or {}
            item_types = response_data.get("item_types", {})
            computer_calls = item_types.get("computer_call", 0)
            if computer_calls > 0:
                tool_calls_total += computer_calls
                tool_calls_by_tool["computer"] = tool_calls_by_tool.get("computer", 0) + computer_calls
    
    # Get timestamps from log file
    log_dir = runner_path / "logs"
    log_file = next(iter(sorted(log_dir.glob("*.log"))), None) if log_dir.exists() else None
    
    start_dt = end_dt = None
    duration_seconds = None
    
    if log_file and log_file.exists():
        try:
            lines = log_file.read_text(encoding="utf-8", errors="ignore").splitlines()
        except Exception:
            lines = []
        timestamp_re = re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3})")
        if lines:
            match = timestamp_re.search(lines[0])
            if match:
                start_dt = _safe_parse_timestamp(match.group(1))
            match = timestamp_re.search(lines[-1])
            if match:
                end_dt = _safe_parse_timestamp(match.group(1))
    
    if start_dt and end_dt:
        duration_seconds = max(0.0, (end_dt - start_dt).total_seconds())
    
    timelapse_str = _format_seconds(duration_seconds) if duration_seconds else None
    
    # Count screenshots
    screenshots_dir = runner_path / "screenshots"
    screenshot_count = len(list(screenshots_dir.glob("*.png"))) if screenshots_dir.exists() else 0
    
    iteration_uuid = (
        verification.get("iteration_uuid")
        or verification.get("iteration_id")
        or (verification.get("iteration") or {}).get("uuid")
    )
    execution_uuid = (
        verification.get("execution_uuid")
        or verification.get("execution_id")
    )
    
    extra_payload: Dict[str, object] = {
        "source": "batch",
        "screenshots_count": screenshot_count,
    }
    if iteration_uuid:
        extra_payload["iteration_uuid"] = iteration_uuid
    if execution_uuid:
        extra_payload["execution_uuid"] = execution_uuid
    
    # Get total_steps from database
    total_steps_value = _get_total_steps_from_db(iteration_uuid) if iteration_uuid else None
    
    # Re-evaluate status with enriched completion_reason
    status = _infer_status(
        status,
        extra=verification,
        status_reason=status_reason,
        completion_reason=completion_reason,
    )
    
    return IterationRecord(
        task_id=task_id,
        iteration=iteration_index or 0,
        runner=runner_name,
        status=status,
        status_reason=status_reason,
        completion_reason=completion_reason,
        duration_seconds=duration_seconds,
        timelapse=timelapse_str,
        file_timelapse_seconds=None,
        tool_calls_total=tool_calls_total,
        tool_calls_by_tool=tool_calls_by_tool,
        unique_tools=list(tool_calls_by_tool.keys()),
        prompt=None,
        model=RUNNER_MODELS.get(runner_name),
        run_id=str(run_id) if run_id else None,
        prompt_id=str(prompt_id) if prompt_id else None,
        start_timestamp=start_dt.isoformat() if start_dt else None,
        end_timestamp=end_dt.isoformat() if end_dt else None,
        iteration_directory=str(runner_path),
        execution_uuid=str(execution_uuid) if execution_uuid else None,
        iteration_uuid=str(iteration_uuid) if iteration_uuid else None,
        verification_comments=None,  # File-based records don't have verification comments
        last_model_response=None,  # File-based records don't have database field
        total_steps=total_steps_value,  # From verification file or database
        extra=extra_payload,
    )


def _load_iteration_record(
    *,
    task_id: str,
    iteration_index: Optional[int],
    runner_name: str,
    runner_path: Path,
) -> Optional[IterationRecord]:
    # Check if this is a batch directory structure (no iteration_meta.json, has batch subdirs)
    batch_subdirs = {'conversation_history', 'task_responses', 'screenshots', 'logs'}
    actual_subdirs = {p.name for p in runner_path.iterdir() if p.is_dir()}
    
    if batch_subdirs.intersection(actual_subdirs) and not (runner_path / "iteration_meta.json").exists():
        # This is a batch directory structure, use batch-specific loading
        return _load_batch_iteration_record(
            task_id=task_id,
            iteration_index=iteration_index,
            runner_name=runner_name,
            runner_path=runner_path,
        )
    
    # Standard execution directory structure with iteration_meta.json
    meta_path = runner_path / "iteration_meta.json"
    meta = _safe_load_json(meta_path)
    if not meta:
        logger.warning("⚠️ Missing iteration_meta.json for %s", runner_path)
        return _fallback_iteration_record(
            task_id=task_id,
            iteration_index=iteration_index,
            runner_name=runner_name,
            runner_path=runner_path,
        )

    extra = meta.get("extra") or {}
    if not isinstance(extra, dict):
        extra = {}

    try:
        tool_usage = meta.get("tool_usage") or {}
        total_tool_calls = int(tool_usage.get("total", 0) or 0)
        by_tool = {
            str(name): int(count)
            for name, count in (tool_usage.get("by_tool") or {}).items()
        }
    except Exception:  # pragma: no cover - safety net
        total_tool_calls = 0
        by_tool = {}

    if not by_tool:
        total_tool_calls, by_tool = _aggregate_tool_calls(runner_path / "api_logs" / "tool_calls.jsonl")

    duration_seconds = _safe_float(meta.get("duration_seconds"))
    if duration_seconds is None:
        duration_seconds = _safe_float(extra.get("file_timelapse_seconds"))
    
    # Fallback: Calculate duration from start/end timestamps if available
    if duration_seconds is None:
        start_ts = meta.get("start_timestamp") or extra.get("start_timestamp")
        end_ts = meta.get("end_timestamp") or extra.get("end_timestamp")
        if start_ts and end_ts:
            try:
                from datetime import datetime
                start_dt = datetime.fromisoformat(start_ts.replace('Z', '+00:00'))
                end_dt = datetime.fromisoformat(end_ts.replace('Z', '+00:00'))
                duration_seconds = (end_dt - start_dt).total_seconds()
                logger.debug(f"Calculated duration from timestamps: {duration_seconds}s for {task_id}")
            except Exception as e:
                logger.debug(f"Could not calculate duration from timestamps: {e}")

    timelapse = extra.get("timelapse")
    if not timelapse and duration_seconds is not None:
        timelapse = _format_seconds(duration_seconds)

    completion_reason = extra.get("completion_reason")
    status_reason = extra.get("status_reason") or extra.get("error")
    model = meta.get("model") or extra.get("model") or RUNNER_MODELS.get(runner_name)
    prompt = meta.get("prompt")
    prompt_id = (
        meta.get("prompt_id")
        or extra.get("prompt_id")
        or (extra.get("backend_verification_results") or {}).get("prompt_id")
    )
    verification_data = _safe_load_json(runner_path / "verification_results.json") or {}
    if not prompt_id:
        prompt_id = verification_data.get("prompt_id")
    iteration_uuid = (
        meta.get("iteration_uuid")
        or extra.get("iteration_uuid")
        or verification_data.get("iteration_uuid")
        or verification_data.get("iteration_id")
        or (verification_data.get("iteration") or {}).get("uuid")
    )
    execution_uuid = (
        meta.get("execution_uuid")
        or extra.get("execution_uuid")
        or verification_data.get("execution_uuid")
        or verification_data.get("execution_id")
    )
    run_id = extra.get("run_id")
    start_timestamp = meta.get("started_at")
    end_timestamp = meta.get("ended_at")
    iteration_directory = meta.get("iteration_directory") or str(runner_path)

    status = _infer_status(
        meta.get("status") or extra.get("status"),
        extra=extra,
        status_reason=status_reason,
        completion_reason=completion_reason,
    )

    if iteration_uuid:
        extra.setdefault("iteration_uuid", iteration_uuid)
    if execution_uuid:
        extra.setdefault("execution_uuid", execution_uuid)
    
    # Get total_steps from database
    total_steps_value = _get_total_steps_from_db(iteration_uuid) if iteration_uuid else None

    return IterationRecord(
        task_id=task_id,
        iteration=iteration_index or 0,
        runner=runner_name,
        status=status,
        status_reason=status_reason,
        completion_reason=str(completion_reason) if completion_reason is not None else None,
        duration_seconds=duration_seconds,
        timelapse=timelapse,
        file_timelapse_seconds=_safe_float(extra.get("file_timelapse_seconds")),
        tool_calls_total=total_tool_calls,
        tool_calls_by_tool=by_tool,
        unique_tools=sorted(by_tool.keys()),
        prompt=str(prompt) if prompt else None,
        prompt_id=str(prompt_id) if prompt_id else None,
        model=str(model) if model else None,
        run_id=str(run_id) if run_id else None,
        start_timestamp=start_timestamp,
        end_timestamp=end_timestamp,
        iteration_directory=iteration_directory,
        execution_uuid=str(execution_uuid) if execution_uuid else None,
        iteration_uuid=str(iteration_uuid) if iteration_uuid else None,
        verification_comments=None,  # File-based records don't have verification comments
        last_model_response=None,  # File-based records don't have database field
        total_steps=total_steps_value,  # From verification file or database
        extra=extra,
    )


# ---------------------------------------------------------------------------
# Fallback loaders for legacy iterations without iteration_meta.json
# ---------------------------------------------------------------------------

def _fallback_iteration_record(
    *,
    task_id: str,
    iteration_index: Optional[int],
    runner_name: str,
    runner_path: Path,
) -> Optional[IterationRecord]:
    runner_key = runner_name.lower()
    if runner_key == "anthropic":
        return _fallback_anthropic_record(task_id, iteration_index, runner_path)
    if runner_key.startswith("openai"):
        return _fallback_openai_record(task_id, iteration_index, runner_path)
    if runner_key == "gemini":
        return _fallback_gemini_record(task_id, iteration_index, runner_path)
    logger.warning("⚠️ No fallback available for runner %s at %s", runner_name, runner_path)
    return None


def _fallback_anthropic_record(
    task_id: str,
    iteration_index: Optional[int],
    runner_path: Path,
) -> Optional[IterationRecord]:
    responses_dir = runner_path / "task_responses"
    response_files = sorted(responses_dir.glob("*.txt")) if responses_dir.exists() else []

    duration_seconds = None
    file_timelapse_seconds = None
    start_timestamp_iso = None
    end_timestamp_iso = None
    if response_files:
        mtimes = [f.stat().st_mtime for f in response_files]
        if mtimes:
            earliest = min(mtimes)
            latest = max(mtimes)
            file_timelapse_seconds = latest - earliest
            duration_seconds = file_timelapse_seconds
            start_timestamp_iso = datetime.fromtimestamp(earliest, tz=timezone.utc).isoformat()
            end_timestamp_iso = datetime.fromtimestamp(latest, tz=timezone.utc).isoformat()

    tool_calls_by_tool: Dict[str, int] = {}
    for path in response_files:
        try:
            for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
                line = line.strip()
                if not line.lower().startswith("tool use:"):
                    continue
                tool_name = line.split(":", 1)[1].strip() or "unknown"
                tool_calls_by_tool[tool_name] = tool_calls_by_tool.get(tool_name, 0) + 1
        except Exception:
            continue

    completion_reason = None
    if response_files:
        latest = max(response_files, key=lambda f: f.stat().st_mtime)
        try:
            completion_reason = latest.read_text(encoding="utf-8", errors="ignore").strip()
        except Exception:
            completion_reason = None

    verification_path = runner_path / "verification_results.json"
    verification = _safe_load_json(verification_path) or {}
    status_reason = verification.get("verification_summary")
    status = _infer_status(
        verification.get("verification_status"),
        extra=verification,
        status_reason=status_reason,
        completion_reason=completion_reason,
    )
    run_id = verification.get("run_id")
    prompt_id = verification.get("prompt_id")

    timelapse_str = _format_seconds(duration_seconds) if duration_seconds else None

    iteration_uuid = (
        verification.get("iteration_uuid")
        or verification.get("iteration_id")
        or (verification.get("iteration") or {}).get("uuid")
    )
    execution_uuid = (
        verification.get("execution_uuid")
        or verification.get("execution_id")
    )

    extra_payload: Dict[str, object] = {"source": "fallback:anthropic"}
    if iteration_uuid:
        extra_payload["iteration_uuid"] = iteration_uuid
    if execution_uuid:
        extra_payload["execution_uuid"] = execution_uuid

    return IterationRecord(
        task_id=task_id,
        iteration=iteration_index or 0,
        runner="anthropic",
        status=status,
        status_reason=status_reason,
        completion_reason=completion_reason,
        duration_seconds=duration_seconds,
        timelapse=timelapse_str,
        file_timelapse_seconds=file_timelapse_seconds,
        tool_calls_total=sum(tool_calls_by_tool.values()),
        tool_calls_by_tool=tool_calls_by_tool,
        unique_tools=sorted(tool_calls_by_tool.keys()),
        prompt=None,
        prompt_id=str(prompt_id) if prompt_id else None,
        model=RUNNER_MODELS.get("anthropic"),
        run_id=str(run_id) if run_id else None,
        start_timestamp=start_timestamp_iso,
        end_timestamp=end_timestamp_iso,
        iteration_directory=str(runner_path),
        execution_uuid=str(execution_uuid) if execution_uuid else None,
        iteration_uuid=str(iteration_uuid) if iteration_uuid else None,
        verification_comments=None,  # File-based records don't have verification comments
        last_model_response=None,  # File-based records don't have database field
        extra=extra_payload,
    )


def _fallback_openai_record(
    task_id: str,
    iteration_index: Optional[int],
    runner_path: Path,
) -> Optional[IterationRecord]:
    # Read verification_results.json for status and run_id
    verification_path = runner_path / "verification_results.json"
    verification = _safe_load_json(verification_path) or {}
    
    status_reason = verification.get("verification_summary")
    completion_reason: Optional[str] = None
    status = _infer_status(
        verification.get("verification_status"),
        extra=verification,
        status_reason=status_reason,
        completion_reason=completion_reason,
    )
    
    # Clean up status_reason if it contains HTML
    if status_reason and "<!doctype" in status_reason.lower():
        # Extract just the error message before the HTML
        if ":" in status_reason:
            status_reason = status_reason.split(":", 1)[0].strip()
        else:
            status_reason = "Invalid verification response"
    
    run_id = verification.get("run_id")
    prompt_id = verification.get("prompt_id")
    # Read task_responses for tool calls only
    task_responses_dir = runner_path / "task_responses"
    tool_calls_total = 0
    tool_calls_by_tool = {}
    
    if task_responses_dir.exists():
        for response_file in sorted(task_responses_dir.glob("*.json")):
            response_data = _safe_load_json(response_file) or {}
            item_types = response_data.get("item_types", {})
            computer_calls = item_types.get("computer_call", 0)
            if computer_calls > 0:
                tool_calls_total += computer_calls
                tool_calls_by_tool["computer"] = tool_calls_by_tool.get("computer", 0) + computer_calls
    
    # Read conversation_history for full completion message (PRIORITY: always check here first)
    conversation_dir = runner_path / "conversation_history"
    if conversation_dir.exists():
        for conv_file in sorted(conversation_dir.glob("*_task_execution_conversation.json")):
            conv_data = _safe_load_json(conv_file) or {}
            messages = conv_data if isinstance(conv_data, list) else conv_data.get("messages", [])
            for msg in reversed(messages):
                if msg.get("role") == "assistant" and msg.get("content"):
                    content = msg["content"]
                    if isinstance(content, str):
                        completion_reason = content  # Full message, no truncation
                    elif isinstance(content, list):
                        for block in content:
                            if isinstance(block, dict) and block.get("type") == "text":
                                completion_reason = block.get("text", "")  # Full message
                                break
                    break
            if completion_reason:
                break
    
    # Re-evaluate status with enriched clues (if we inferred additional details)
    status = _infer_status(
        status,
        extra=verification,
        status_reason=status_reason,
        completion_reason=completion_reason,
    )

    # Count screenshots
    screenshots_dir = runner_path / "screenshots"
    screenshot_count = len(list(screenshots_dir.glob("*.png"))) if screenshots_dir.exists() else 0
    
    # Get timestamps from log file
    log_dir = runner_path / "logs"
    log_file = next(iter(sorted(log_dir.glob("*.log"))), None) if log_dir.exists() else None
    
    start_dt = end_dt = None
    if log_file and log_file.exists():
        try:
            lines = log_file.read_text(encoding="utf-8", errors="ignore").splitlines()
        except Exception:
            lines = []
        timestamp_re = re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3})")
        if lines:
            match = timestamp_re.search(lines[0])
            if match:
                start_dt = _safe_parse_timestamp(match.group(1))
            match = timestamp_re.search(lines[-1])
            if match:
                end_dt = _safe_parse_timestamp(match.group(1))

    duration_seconds = None
    if start_dt and end_dt:
        duration_seconds = max(0.0, (end_dt - start_dt).total_seconds())

    timelapse_str = _format_seconds(duration_seconds) if duration_seconds else None

    iteration_uuid = (
        verification.get("iteration_uuid")
        or verification.get("iteration_id")
        or (verification.get("iteration") or {}).get("uuid")
    )
    execution_uuid = (
        verification.get("execution_uuid")
        or verification.get("execution_id")
    )

    extra_payload: Dict[str, object] = {
        "source": "fallback:openai",
        "screenshots_count": screenshot_count,
    }
    if iteration_uuid:
        extra_payload["iteration_uuid"] = iteration_uuid
    if execution_uuid:
        extra_payload["execution_uuid"] = execution_uuid

    return IterationRecord(
        task_id=task_id,
        iteration=iteration_index or 0,
        runner="openai",
        status=status,
        status_reason=status_reason,
        completion_reason=completion_reason,
        duration_seconds=duration_seconds,
        timelapse=timelapse_str,
        file_timelapse_seconds=None,
        tool_calls_total=tool_calls_total,
        tool_calls_by_tool=tool_calls_by_tool,
        unique_tools=list(tool_calls_by_tool.keys()),
        prompt=None,
        model=RUNNER_MODELS.get("openai"),
        run_id=run_id,
        prompt_id=str(prompt_id) if prompt_id else None,
        start_timestamp=start_dt.isoformat() if start_dt else None,
        end_timestamp=end_dt.isoformat() if end_dt else None,
        iteration_directory=str(runner_path),
        execution_uuid=str(execution_uuid) if execution_uuid else None,
        iteration_uuid=str(iteration_uuid) if iteration_uuid else None,
        verification_comments=None,  # File-based records don't have verification comments
        last_model_response=None,  # File-based records don't have database field
        extra=extra_payload,
    )


def _fallback_gemini_record(
    task_id: str,
    iteration_index: Optional[int],
    runner_path: Path,
) -> Optional[IterationRecord]:
    # Read verification_results.json for status and run_id
    verification_path = runner_path / "verification_results.json"
    verification = _safe_load_json(verification_path) or {}
    
    status_reason = verification.get("verification_summary")
    completion_reason: Optional[str] = None
    status = _infer_status(
        verification.get("verification_status"),
        extra=verification,
        status_reason=status_reason,
        completion_reason=completion_reason,
    )
    
    # Clean up status_reason if it contains HTML
    if status_reason and "<!doctype" in status_reason.lower():
        # Extract just the error message before the HTML
        if ":" in status_reason:
            status_reason = status_reason.split(":", 1)[0].strip()
        else:
            status_reason = "Invalid verification response"
    
    run_id = verification.get("run_id")
    prompt_id = verification.get("prompt_id")
    
    # Read task_responses for tool calls only
    task_responses_dir = runner_path / "task_responses"
    tool_calls_total = 0
    tool_calls_by_tool = {}
    
    if task_responses_dir.exists():
        for response_file in sorted(task_responses_dir.glob("*.json")):
            response_data = _safe_load_json(response_file) or {}
            item_types = response_data.get("item_types", {})
            computer_calls = item_types.get("computer_call", 0)
            if computer_calls > 0:
                tool_calls_total += computer_calls
                tool_calls_by_tool["computer"] = tool_calls_by_tool.get("computer", 0) + computer_calls
    
    # Read conversation_history for full completion message (PRIORITY: always check here first)
    conversation_dir = runner_path / "conversation_history"
    if conversation_dir.exists():
        for conv_file in sorted(conversation_dir.glob("*_task_execution_conversation.json")):
            conv_data = _safe_load_json(conv_file) or {}
            messages = conv_data if isinstance(conv_data, list) else conv_data.get("messages", [])
            for msg in reversed(messages):
                if msg.get("role") == "assistant" and msg.get("content"):
                    content = msg["content"]
                    if isinstance(content, str):
                        completion_reason = content  # Full message, no truncation
                    elif isinstance(content, list):
                        for block in content:
                            if isinstance(block, dict) and block.get("type") == "text":
                                completion_reason = block.get("text", "")  # Full message
                                break
                    break
            if completion_reason:
                break
    
    # Re-evaluate status with enriched clues (if we inferred additional details)
    status = _infer_status(
        status,
        extra=verification,
        status_reason=status_reason,
        completion_reason=completion_reason,
    )

    # Count screenshots
    screenshots_dir = runner_path / "screenshots"
    screenshot_count = len(list(screenshots_dir.glob("*.png"))) if screenshots_dir.exists() else 0
    
    # Get timestamps from log file
    log_dir = runner_path / "logs"
    log_file = next(iter(sorted(log_dir.glob("*.log"))), None) if log_dir.exists() else None
    
    start_dt = end_dt = None
    if log_file and log_file.exists():
        try:
            lines = log_file.read_text(encoding="utf-8", errors="ignore").splitlines()
        except Exception:
            lines = []
        timestamp_re = re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3})")
        if lines:
            match = timestamp_re.search(lines[0])
            if match:
                start_dt = _safe_parse_timestamp(match.group(1))
            match = timestamp_re.search(lines[-1])
            if match:
                end_dt = _safe_parse_timestamp(match.group(1))

    duration_seconds = None
    if start_dt and end_dt:
        duration_seconds = max(0.0, (end_dt - start_dt).total_seconds())

    timelapse_str = _format_seconds(duration_seconds) if duration_seconds else None

    iteration_uuid = (
        verification.get("iteration_uuid")
        or verification.get("iteration_id")
        or (verification.get("iteration") or {}).get("uuid")
    )
    execution_uuid = (
        verification.get("execution_uuid")
        or verification.get("execution_id")
    )

    extra_payload: Dict[str, object] = {
        "source": "fallback:gemini",
        "screenshots_count": screenshot_count,
    }
    if iteration_uuid:
        extra_payload["iteration_uuid"] = iteration_uuid
    if execution_uuid:
        extra_payload["execution_uuid"] = execution_uuid

    return IterationRecord(
        task_id=task_id,
        iteration=iteration_index or 0,
        runner="gemini",
        status=status,
        status_reason=status_reason,
        completion_reason=completion_reason,
        duration_seconds=duration_seconds,
        timelapse=timelapse_str,
        file_timelapse_seconds=None,
        tool_calls_total=tool_calls_total,
        tool_calls_by_tool=tool_calls_by_tool,
        unique_tools=list(tool_calls_by_tool.keys()),
        prompt=None,
        model=RUNNER_MODELS.get("gemini"),
        run_id=run_id,
        prompt_id=str(prompt_id) if prompt_id else None,
        start_timestamp=start_dt.isoformat() if start_dt else None,
        end_timestamp=end_dt.isoformat() if end_dt else None,
        iteration_directory=str(runner_path),
        execution_uuid=str(execution_uuid) if execution_uuid else None,
        iteration_uuid=str(iteration_uuid) if iteration_uuid else None,
        verification_comments=None,  # File-based records don't have verification comments
        last_model_response=None,  # File-based records don't have database field
        extra=extra_payload,
    )


# GPT-4o Advanced fallback function removed - temporarily suspended
def _compute_task_timings(records: List[IterationRecord]) -> Dict[str, Optional[float]]:
    if not records:
        return {
            "total_seconds": None,
            "average_iteration_seconds": None,
            "start_timestamp": None,
            "end_timestamp": None,
        }

    iteration_durations = [r.duration_seconds for r in records if r.duration_seconds is not None]
    earliest = [r.start_timestamp for r in records if r.start_timestamp]
    latest = [r.end_timestamp for r in records if r.end_timestamp]

    total_seconds = sum(iteration_durations) if iteration_durations else None
    avg_seconds = (total_seconds / len(iteration_durations)) if iteration_durations else None

    start_ts = min(earliest) if earliest else None
    end_ts = max(latest) if latest else None

    return {
        "total_seconds": total_seconds,
        "average_iteration_seconds": avg_seconds,
        "start_timestamp": start_ts,
        "end_timestamp": end_ts,
    }


def _compute_median_steps(records: List[IterationRecord]) -> Optional[int]:
    """
    Compute median steps from a list of iteration records.
    Only considers records with total_steps data.
    Returns None if no steps data is available.
    """
    from statistics import median
    
    steps_list = []
    for record in records:
        if record.total_steps is not None:
            try:
                steps_list.append(int(record.total_steps))
            except (ValueError, TypeError):
                continue
    
    if not steps_list:
        return None
    
    return int(median(steps_list))


def _build_summary(
    records: List[IterationRecord],
    *,
    filtered_iterations: Optional[List[IterationRecord]] = None,
) -> Tuple[List[Dict[str, object]], Dict[str, Dict[str, List[IterationRecord]]]]:
    task_map: Dict[str, Dict[str, List[IterationRecord]]] = defaultdict(lambda: defaultdict(list))

    for record in records:
        task_map[record.task_id][record.runner].append(record)

    # Ensure deterministic ordering of iterations per runner
    for runner_map in task_map.values():
        for runner_records in runner_map.values():
            runner_records.sort(key=lambda r: r.iteration)

    summary_rows: List[Dict[str, object]] = []
    for task_id, runner_map in task_map.items():
        prompt = _select_prompt(runner_map)
        prompt_id = _select_prompt_id(runner_map) or task_id

        row = {
            "Prompt ID": prompt_id or "",
            "Task": task_id,
            "Prompt": prompt or "",
        }

        # Compute metrics using all available iterations across runners
        flattened_records: List[IterationRecord] = []
        for runner_records in runner_map.values():
            flattened_records.extend(runner_records)

        timings = _compute_task_timings(flattened_records)
        row["Total Time Seconds"] = timings["total_seconds"]
        row["Task Start Time"] = timings["start_timestamp"]
        row["Task End Time"] = timings["end_timestamp"]
        row["Average Iteration Time Minutes"] = (
            round(timings["average_iteration_seconds"] / 60.0, 2)
            if timings["average_iteration_seconds"] is not None
            else None
        )

        # Only include models that actually have data in this batch
        for runner_key, label in MODEL_ORDER:
            records_for_runner = runner_map.get(runner_key, [])
            if records_for_runner:  # Only add columns for models with data
                stats = _aggregate_runner_stats(records_for_runner)
                row[f"{label} Breaking"] = _format_breaking_string(stats)
                # Store individual model difficulty for styling
                row[f"{label} Difficulty"] = _determine_model_difficulty_from_stats(stats)
                # Compute and store median steps for this model
                row[f"{label} Median Steps"] = _compute_median_steps(records_for_runner)

        row["Difficulty"] = _difficulty_from_runner_stats(runner_map)
        summary_rows.append(row)

    summary_rows.sort(key=lambda row: row["Task"])
    return summary_rows, task_map


# ---------------------------------------------------------------------------
# Workbook writer
# ---------------------------------------------------------------------------

def _determine_model_difficulty_from_stats(stats: Dict[str, object]) -> str:
    """
    Determine difficulty for a single model based on its stats.
    Uses new two-criteria system (Pass Rate + Steps) when steps are available.
    """
    from statistics import median
    
    total = stats.get("total", 0)
    if not total:
        return "unknown"  # No data available
    
    pass_count = stats.get("pass_count", 0)
    success_rate = (pass_count / total) * 100
    
    # Check if we have step data available
    records = stats.get("records", [])
    steps_list = []
    for rec in records:
        if hasattr(rec, 'total_steps') and rec.total_steps is not None:
            steps_list.append(rec.total_steps)
    
    # If we have steps data, use the new two-criteria system
    if steps_list:
        median_steps = int(median(steps_list))
        
        # NEW two-criteria logic (matching task sheet and definitions):
        # Hard: success_rate < 40% OR median_steps > 50
        if success_rate < 40.0 or median_steps > 50:
            return "hard"
        
        # Easy: success_rate == 100% AND median_steps <= 25
        if success_rate == 100.0 and median_steps <= 25:
            return "easy"
        
        # Medium: Everything else
        return "medium"
    
    # Legacy system without steps (old behavior preserved)
    if success_rate >= 100:
        return "easy"  # 100% success
    elif success_rate >= 40:
        return "medium"  # 40-99% success
    else:
        return "hard"  # <40% success


def _determine_model_difficulty(model_result: str) -> str:
    """
    Determine difficulty for a single model result.
    Note: The result format is "Breaking: Yes/No, X/Y"
    - "Yes, X/Y" means model IS breaking (failing) X out of Y times
    - "No, 0/Y" means model is NOT breaking (passing all Y times)
    """
    result_lower = str(model_result).lower()
    
    if result_lower.startswith("yes"):
        # "Yes, X/Y" means model IS breaking X out of Y times
        import re
        match = re.search(r'(\d+)/(\d+)', result_lower)
        if match:
            break_count = int(match.group(1))  # Number of failures
            total_count = int(match.group(2))
            failure_rate = (break_count / total_count) * 100
            
            # High failure rate = hard task
            if failure_rate >= 60:
                return "hard"  # ≥60% failures = hard task
            elif failure_rate >= 20:
                return "medium"  # 20-59% failures = medium task
            else:
                return "easy"  # <20% failures = easy task
        else:
            return "hard"  # Default for "Yes" without ratio means model is breaking
    elif result_lower.startswith("no"):
        return "easy"  # Model is NOT breaking (passing all) = easy task
    else:
        return "unknown"


def _determine_difficulty_from_results(row: Dict[str, object]) -> str:
    """
    Determine overall difficulty for the row (used for Difficulty column).
    Note: The column is "Breaking" so "Yes" means model IS breaking (failing).
    """
    # Get model results
    claude_result = str(row.get("Claude Sonnet 4 Breaking", "")).lower()
    openai_result = str(row.get("OpenAI Computer Use Preview Breaking", "")).lower()
    gemini_result = str(row.get("Google Gemini Computer Use Breaking", "")).lower()
    
    # Count breaking results (models that are failing - start with "yes")
    breaking_count = 0
    if claude_result.startswith("yes"):
        breaking_count += 1
    if openai_result.startswith("yes"):
        breaking_count += 1
    if gemini_result.startswith("yes"):
        breaking_count += 1
    
    # Determine difficulty based on how many models are breaking
    # More models breaking = harder task
    if breaking_count == 0:
        return "easy"  # No models breaking (all passing) = easy task
    elif breaking_count == 1:
        return "medium"  # One model breaking = medium task
    elif breaking_count == 2:
        return "medium"  # Two models breaking = medium task
    else:
        return "hard"  # All models breaking (all failing) = hard task


def _get_difficulty_colors(difficulty: str) -> Dict[str, str]:
    """Get colors for difficulty-based styling matching the difficulty definitions table exactly."""
    if difficulty == "hard":
        return {
            'bg': 'FFC7CE',  # Light red background - matches difficulty definitions table
            'text': '000000',  # Black text for readability
            'border': 'FFC7CE'  # Same as background
        }
    elif difficulty == "medium":
        return {
            'bg': 'FFEB9C',  # Light yellow background - matches difficulty definitions table
            'text': '000000',  # Black text for readability
            'border': 'FFEB9C'  # Same as background
        }
    elif difficulty == "easy":
        return {
            'bg': 'C6EFCE',  # Light green background - matches difficulty definitions table
            'text': '000000',  # Black text for readability
            'border': 'C6EFCE'  # Same as background
        }
    else:
        return {
            'bg': 'FFFFFF',  # White background for unknown/empty
            'text': '000000',  # Black text
            'border': 'FFFFFF'  # Same as background
        }


def _get_model_colors(runner_key: str) -> Dict[str, str]:
    """Get model-specific colors based on runner key."""
    runner_lower = runner_key.lower()
    if 'gpt' in runner_lower or 'openai' in runner_lower:
        return {
            'header_bg': '1976D2',  # Deep blue
            'header_text': 'FFFFFF',  # White
            'data_bg': 'E3F2FD',    # Light blue
            'data_text': '000000'    # Black
        }
    elif 'claude' in runner_lower or 'anthropic' in runner_lower:
        return {
            'header_bg': '7B1FA2',  # Dark purple
            'header_text': 'FFFFFF',  # White
            'data_bg': 'F3E5F5',    # Light purple
            'data_text': '000000'    # Black
        }
    elif 'gemini' in runner_lower or 'google' in runner_lower:
        return {
            'header_bg': '5D4037',  # Dark brown
            'header_text': 'FFFFFF',  # White
            'data_bg': 'EFEBE9',    # Light brown/cream
            'data_text': '000000'    # Black
        }
    else:
        # Default colors for unknown models
        return {
            'header_bg': '424242',  # Dark gray
            'header_text': 'FFFFFF',  # White
            'data_bg': 'F5F5F5',    # Light gray
            'data_text': '000000'    # Black
        }


def _write_workbook(
    *,
    summary_rows: List[Dict[str, object]],
    iterations: List[IterationRecord],
    task_rows: Dict[str, Dict[str, List[IterationRecord]]],
    workbook_path: Path,
    total_iterations: int | None = None,
    batch_insights: Dict[str, str] = None,
) -> None:
    """
    Write the complete Excel workbook with summary and task sheets.
    
    Column Width Consistency Strategy:
    - Summary sheet: Uses include_steps flag (global - based on ANY iteration having steps)
        - include_steps=False: 1 column per model (Breaking only)
        - include_steps=True: 2 columns per model (Breaking + Median Steps)
    
    - Task sheets: Each task checks its own iterations for steps data (per-task basis)
        - has_steps_data=False: 5 columns per model (no Steps column)
        - has_steps_data=True: 6 columns per model (includes Steps column)
    
    This ensures:
    1. Summary sheet is consistent across all tasks (global decision)
    2. Each task sheet is consistent within itself (local decision)
    3. Old reports without steps show no step columns anywhere
    4. New reports with steps show step columns and charts
    """
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    # Determine if step data is available anywhere to enhance difficulty criteria text
    def _has_steps(rec: IterationRecord) -> bool:
        return rec.total_steps is not None

    include_steps_in_definitions = any(_has_steps(r) for r in (iterations or []))

    _populate_summary_sheet(summary_ws, summary_rows, total_iterations, include_steps=include_steps_in_definitions)

    for task_id, runner_map in sorted(task_rows.items()):
        sheet_name = _safe_sheet_name(wb, task_id)
        task_ws = wb.create_sheet(sheet_name)
        _populate_task_sheet(task_ws, task_id, runner_map)

    wb.save(workbook_path)


def _populate_summary_sheet(ws, rows: List[Dict[str, object]], total_iterations: int | None = None, *, include_steps: bool = False) -> None:
    """
    Populate the summary sheet with task results.
    
    Column structure depends on include_steps flag:
    
    When include_steps=False (old reports without step data):
        Col A: Prompt ID
        Col B: Prompt
        Col C: Model1 Breaking
        Col D: Model2 Breaking
        Col E: Model3 Breaking
        ... (one column per model)
    
    When include_steps=True (new reports with step data):
        Col A: Prompt ID
        Col B: Prompt
        Col C: Model1 Breaking
        Col D: Model1 Median Steps
        Col E: Model2 Breaking
        Col F: Model2 Median Steps
        Col G: Model3 Breaking
        Col H: Model3 Median Steps
        ... (two columns per model)
    
    This ensures backward compatibility with old reports while supporting new step data.
    """
    # Determine which model columns actually have data
    model_columns = []
    if rows:
        # Check which model columns exist in the data
        for model_key, model_label in MODEL_ORDER:
            column_name = f"{model_label} Breaking"
            if any(column_name in row for row in rows):
                model_columns.append((model_key, model_label, column_name))
    
    # Build headers dynamically - include Breaking and optionally Median Steps columns
    headers = ["Prompt ID", "Prompt"]
    for _, model_label, _ in model_columns:
        headers.append(f"{model_label} Breaking")
        if include_steps:
            headers.append(f"{model_label} Median Steps")
    
    ws.append(headers)

    # Apply header styling
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, 2):  # Start from row 2 (after header)
        # Add row data dynamically
        row_data = [
            _sanitize_for_excel(str(row.get("Prompt ID", ""))),
            _sanitize_for_excel(str(row.get("Prompt", ""))),
        ]
        # Add model columns dynamically - Breaking and optionally Median Steps
        for _, model_label, _ in model_columns:
            row_data.append(_sanitize_for_excel(str(row.get(f"{model_label} Breaking", ""))))
            # Only add median steps if include_steps is True
            if include_steps:
                # Add median steps value as NUMBER (not string) for charts to work
                median_steps = row.get(f"{model_label} Median Steps")
                if median_steps is not None:
                    try:
                        row_data.append(int(median_steps))  # Keep as integer for chart
                    except (ValueError, TypeError):
                        row_data.append(0)  # Fallback to 0 if conversion fails
                else:
                    row_data.append(0)  # Use 0 instead of empty string
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply individual model difficulty-based colors to model Breaking columns
            # Column structure depends on include_steps:
            # - Without steps: Prompt ID (1), Prompt (2), Breaking1 (3), Breaking2 (4), ...
            # - With steps: Prompt ID (1), Prompt (2), Breaking1 (3), MedianSteps1 (4), Breaking2 (5), MedianSteps2 (6), ...
            if col_idx > 2:  # After Prompt columns
                col_offset = col_idx - 3  # 0-based offset from first model column
                
                # Determine if this is a Breaking column based on include_steps
                is_breaking_column = False
                if include_steps:
                    # With steps: Breaking columns are at even offsets (0, 2, 4, ...)
                    is_breaking_column = (col_offset % 2 == 0)
                    model_idx = col_offset // 2
                else:
                    # Without steps: All columns after Prompt are Breaking columns
                    is_breaking_column = True
                    model_idx = col_offset
                
                if is_breaking_column and model_idx < len(model_columns):
                    _, model_label, _ = model_columns[model_idx]
                    difficulty_column_name = f"{model_label} Difficulty"
                    model_difficulty = row.get(difficulty_column_name, "unknown")
                    
                    # Get colors for this specific model's difficulty
                    model_difficulty_colors = _get_difficulty_colors(model_difficulty)
                    
                    # Only apply background color if there's actual data (not empty or "No data")
                    if str(value).strip() and str(value).strip() not in ["No data", ""]:
                        cell.fill = PatternFill(
                            start_color=model_difficulty_colors['bg'], 
                            end_color=model_difficulty_colors['bg'], 
                            fill_type="solid"
                        )
                        cell.font = Font(color=model_difficulty_colors['text'])
                    else:
                        # Keep empty cells white with no special styling
                        cell.fill = PatternFill(
                            start_color='FFFFFF', 
                            end_color='FFFFFF', 
                            fill_type="solid"
                        )
                        cell.font = Font(color='000000')

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Set column widths dynamically - Breaking and optionally Median Steps for each model
    column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    
    col_idx = 0
    ws.column_dimensions[column_letters[col_idx]].width = 28   # Prompt ID
    col_idx += 1
    ws.column_dimensions[column_letters[col_idx]].width = 120  # Prompt
    col_idx += 1
    
    # Each model has Breaking column (35) and optionally Median Steps (35)
    BREAKING_COLUMN_WIDTH = 35
    MEDIAN_STEPS_COLUMN_WIDTH = 35  # Same width as Breaking column
    for _ in model_columns:
        ws.column_dimensions[column_letters[col_idx]].width = BREAKING_COLUMN_WIDTH
        col_idx += 1
        if include_steps:
            ws.column_dimensions[column_letters[col_idx]].width = MEDIAN_STEPS_COLUMN_WIDTH
            col_idx += 1


    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap_alignment
        ws.row_dimensions[row[0].row].height = 80
    
    # Add Difficulty Definitions table 3 rows below the last entry
    definitions_start_row = ws.max_row + 3
    _add_difficulty_definitions_table(ws, definitions_start_row, total_iterations, include_steps=include_steps)
    
    # Re-apply summary sheet column widths after Difficulty Definitions table
    # (The table may have overridden some column widths)
    col_idx = 0
    ws.column_dimensions[column_letters[col_idx]].width = 28   # Prompt ID
    col_idx += 1
    ws.column_dimensions[column_letters[col_idx]].width = 120  # Prompt
    col_idx += 1
    
    # Re-apply model column widths (must match above)
    for _ in model_columns:
        ws.column_dimensions[column_letters[col_idx]].width = BREAKING_COLUMN_WIDTH
        col_idx += 1
        if include_steps:
            ws.column_dimensions[column_letters[col_idx]].width = MEDIAN_STEPS_COLUMN_WIDTH
            col_idx += 1
    
    # Add Median Steps chart only if steps are available
    if include_steps:
        _add_median_steps_chart(ws, rows, model_columns, definitions_start_row)



def _add_median_steps_chart(ws, rows: List[Dict[str, object]], model_columns: List[Tuple[str, str, str]], definitions_start_row: int) -> None:
    """
    Add a SIMPLE bar chart showing median steps per task.
    Chart reads directly from the existing summary table.
    
    CRITICAL: This function assumes the summary sheet was created WITH include_steps=True,
    which means the column structure is:
    Col A=Prompt ID, B=Prompt, C=Breaking1, D=MedianSteps1, E=Breaking2, F=MedianSteps2, etc.
    
    This function should ONLY be called when include_steps=True.
    """
    try:
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        
        if not rows or not model_columns:
            logger.info("Skipping chart - no data")
            return
        
        # Validate that we have actual numeric data for charts
        has_valid_data = False
        for row in rows:
            for _, model_label, _ in model_columns:
                median_steps = row.get(f"{model_label} Median Steps")
                if median_steps is not None and median_steps != 0:
                    has_valid_data = True
                    break
            if has_valid_data:
                break
        
        if not has_valid_data:
            logger.info("Skipping chart - no valid median steps data")
            return
        
        # Simple bar chart setup
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Median Steps by Task and Model"
        
        # Use fixed width for consistent spacing across all reports
        num_rows = len(rows)
        
        # Chart dimensions - FIXED width for consistency
        chart.height = 15
        chart.width = 30  # Fixed 30 cm width for all reports
        
        # Axis titles
        chart.y_axis.title = "Median Steps"
        chart.x_axis.title = "Task ID"
        
        # No gridlines
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None
        
        # Calculate max value for Y-axis to ensure proper scaling and visibility
        max_steps = 0
        for row in rows:
            for _, model_label, _ in model_columns:
                median_steps = row.get(f"{model_label} Median Steps")
                if median_steps is not None:
                    max_steps = max(max_steps, int(median_steps))
        
        # Set Y-axis scaling for better number visibility
        if max_steps > 0:
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = int(max_steps * 1.2)  # 20% headroom
        
        # CRITICAL: Explicitly prevent axes from being deleted (common openpyxl issue)
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        
        # Configure X-axis to show all labels clearly
        # tickLblSkip must be >= 1 (1 = show every label, 2 = skip every other, etc.)
        chart.x_axis.tickLblSkip = 1  # 1 = show ALL labels (minimum valid value)
        
        # NO data labels on bars (removed to avoid clutter)
        # Users can see exact values in the data table above
        
        # Add each model's data as a series
        # IMPORTANT: This function is ONLY called when include_steps=True
        # Data structure with steps: Col A=Prompt ID, B=Prompt, C=Break1, D=Steps1, E=Break2, F=Steps2, G=Break3, H=Steps3
        # Without steps, the structure would be different, but this chart won't be rendered
        
        # For each model, add their median steps column
        for idx, (_, model_label, _) in enumerate(model_columns):
            # Column positions when include_steps=True:
            # Model 0: Breaking=C(3), Steps=D(4)
            # Model 1: Breaking=E(5), Steps=F(6)
            # Model 2: Breaking=G(7), Steps=H(8)
            col_num = 4 + (idx * 2)  # Column D=4, F=6, H=8 (Median Steps columns)
            
            # Data reference: from row 1 (header) to last data row
            data = Reference(ws, min_col=col_num, min_row=1, max_row=num_rows + 1)
            chart.add_data(data, titles_from_data=True)
        
        # Categories (X-axis labels) from Prompt ID column
        cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows + 1)
        chart.set_categories(cats)
        
        # Legend position - top right
        chart.legend.position = "tr"  # Top right
        chart.legend.overlay = False  # Don't overlay the chart
        
        # Adjust plot area to prevent labels from overlapping with chart edges
        # This creates internal margins within the chart box
        # Use consistent spacing for all reports (fixed width ensures no overlap)
        from openpyxl.chart.layout import Layout, ManualLayout
        
        chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.1,   # 10% left margin - balanced spacing
                y=0.1,   # Start 10% from top (space for title/legend)
                w=0.85,  # 85% width - good use of space
                h=0.55   # Height 55% of chart box (more space for X-axis labels + title at bottom)
            )
        )
        
        # Place chart
        ws.add_chart(chart, f"C{definitions_start_row + 8}")
        
    except Exception as e:
        logger.error(f"Chart error: {e}")
        import traceback
        logger.error(traceback.format_exc())


def _add_difficulty_definitions_table(ws, start_row: int, total_iterations: int | None = None, *, include_steps: bool = False) -> None:
    """Add the Difficulty Definitions table to the summary sheet.

    If total_iterations is provided, use Pass@N based on that value; otherwise, infer N
    from the data range (fallback to 8 if not inferable).
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    
    # Title row - "Difficulty Definitions" spanning 3 columns starting from column C (one column right of Prompt column)
    title_cell = ws.cell(row=start_row, column=3, value="Difficulty Definitions")
    title_cell.font = Font(bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=5)
    
    # Header row
    header_row = start_row + 1
    headers = ["Difficulty", "Criteria", "Description"]
    for col_idx, header in enumerate(headers, 3):  # Start from column C
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Determine N for Pass@N definitions
    N = int(total_iterations) if isinstance(total_iterations, int) and total_iterations > 0 else 8

    # Compute integer ranges for descriptions
    import math
    pass_min_medium = math.floor(0.4 * N) + 1  # strictly greater than 40%
    pass_max_medium = max(N - 1, 0)            # strictly less than 100%
    fail_min_medium = 1 if N > 1 else 0
    fail_max_medium = max(N - pass_min_medium, 0)
    pass_max_hard = math.floor(0.4 * N)
    fail_min_hard = math.ceil(0.6 * N)

    # Data rows - handle with and without steps
    difficulty_data = []
    
    # Easy difficulty - always present
    if include_steps:
        # Two-criteria system: Pass Rate AND Steps (NEW thresholds: 25/50/50+)
        if N == 1:
            criteria = f"Pass@{N} = 100% AND Median Steps ≤ 25"
            description = "Prompt should pass the single iteration AND Median Steps should be ≤ 25"
        else:
            criteria = f"Pass@{N} = 100% AND Median Steps ≤ 25"
            description = f"Prompt should pass {N} times (0 failures) AND Median Steps should be ≤ 25"
    else:
        if N == 1:
            criteria = f"Pass@{N} = 100%"
            description = "Prompt should pass the single iteration"
        else:
            criteria = f"Pass@{N} = 100%"
            description = f"Prompt should fail 0 times and pass {N} times"
    difficulty_data.append({
        "difficulty": "Easy",
        "criteria": criteria,
        "description": description,
        "fill_color": "C6EFCE"  # Light green
    })
    
    # Add Medium difficulty
    if include_steps:
        # WITH steps: Always show Medium for all N (including N=1)
        if N == 1:
            criteria = "(Pass@1 < 100% AND Pass@1 ≥ 50%) OR (25 < Median Steps ≤ 50)"
            description = "Prompt fails single iteration but could pass OR Median Steps between 26-50"
        elif N == 2:
            criteria = f"(50% ≤ Pass@{N} < 100%) OR (25 < Median Steps ≤ 50)"
            description = f"Prompt should pass 1 out of 2 times OR Median Steps between 26-50"
        else:
            criteria = f"(40% < Pass@{N} < 100%) OR (25 < Median Steps ≤ 50)"
            description = (
                f"Prompt should pass between {pass_min_medium}-{pass_max_medium} out of {N} times OR "
                f"Median Steps between 26-50"
            )
        difficulty_data.append({
            "difficulty": "Medium", 
            "criteria": criteria,
            "description": description,
            "fill_color": "FFEB9C"  # Light yellow
        })
    else:
        # WITHOUT steps: Only show Medium for N >= 2 (legacy behavior)
        if N >= 2:
            if N == 2:
                criteria = f"Pass@{N} ≥ 50%"
                description = f"Prompt should pass 1 out of 2 times and fail 1 out of 2 times"
            else:
                criteria = f"40% < Pass@{N} < 100%"
                description = (
                    f"Prompt should fail between {fail_min_medium}-{fail_max_medium} out of {N} times and "
                    f"pass between {pass_min_medium}-{pass_max_medium} out of {N} times"
                )
            difficulty_data.append({
                "difficulty": "Medium", 
                "criteria": criteria,
                "description": description,
                "fill_color": "FFEB9C"  # Light yellow
            })
    
    # Add Hard difficulty - always present
    if include_steps:
        # Two-criteria system with NEW thresholds: (Pass Rate < threshold) OR (Steps > 50)
        if N == 1:
            criteria = "Pass@1 < 100% OR Median Steps > 50"
            description = "Prompt should fail the single iteration OR Median Steps should be > 50"
        elif N == 2:
            criteria = "Pass@2 < 50% OR Median Steps > 50"
            description = "Prompt should fail both iterations OR Median Steps should be > 50"
        else:
            criteria = f"Pass@{N} < 40% OR Median Steps > 50"
            description = (
                f"Prompt should pass between 0-{pass_max_hard} out of {N} times OR "
                f"Median Steps should be > 50"
            )
    else:
        # Old system without steps - unchanged
        if N == 1:
            criteria = "Pass@1 < 100%"
            description = f"Prompt should fail the single iteration and pass 0 times"
        elif N == 2:
            criteria = f"Pass@{N} < 40%"
            description = f"Prompt should fail both iterations and pass 0 times"
        else:
            criteria = f"Pass@{N} < 40%"
            description = (
                f"Prompt should fail between {fail_min_hard}-{N} out of {N} times and "
                f"pass between 0-{pass_max_hard} out of {N} times"
            )
    difficulty_data.append({
        "difficulty": "Hard",
        "criteria": criteria, 
        "description": description,
        "fill_color": "FFC7CE"  # Light red
    })
    
    for row_idx, data in enumerate(difficulty_data, header_row + 1):
        # Difficulty column with colored background
        difficulty_cell = ws.cell(row=row_idx, column=3, value=data["difficulty"])
        difficulty_cell.fill = PatternFill(start_color=data["fill_color"], end_color=data["fill_color"], fill_type="solid")
        difficulty_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Criteria column with wrapped text to prevent breaking
        criteria_cell = ws.cell(row=row_idx, column=4, value=data["criteria"])
        criteria_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Description column with wrapped text
        description_cell = ws.cell(row=row_idx, column=5, value=data["description"])
        description_cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Set row height for better text wrapping
        ws.row_dimensions[row_idx].height = 60
    
    # Set explicit column widths for difficulty criteria table for better readability
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(3)].width = 20   # Difficulty column (C)
    ws.column_dimensions[get_column_letter(4)].width = 40   # Criteria column (D) - minimum 40 for readability
    ws.column_dimensions[get_column_letter(5)].width = 40   # Description column (E) - minimum 40 for readability


def _populate_task_sheet(ws, task_id: str, runner_map: Dict[str, List[IterationRecord]]) -> None:
    """
    Populate a task-specific sheet with iteration details.
    
    Column structure per model depends on whether steps data is available:
    
    When has_steps_data=False (no steps in any iteration):
        - Model Response
        - Pass/Fail
        - Time Taken
        - Iteration Link
        - Comments
        (5 columns per model)
    
    When has_steps_data=True (at least one iteration has steps):
        - Model Response
        - Pass/Fail
        - Time Taken
        - Steps
        - Iteration Link
        - Comments
        (6 columns per model)
    
    This ensures consistency: either ALL tasks show steps or NONE do (per task basis).
    """
    base_execution_url = settings.frontend_base_url

    # Check if any iteration in this task has steps data
    has_steps_data = False
    for records in runner_map.values():
        for record in records:
            if record.total_steps is not None:
                has_steps_data = True
                break
        if has_steps_data:
            break

    # Build column headers based on steps availability
    model_column_headers = [
        "Model Response",
        "Pass/Fail",
        "Time Taken",
    ]
    if has_steps_data:
        model_column_headers.append("Steps")
    model_column_headers.extend([
        "Iteration Link",
        "Comments",
    ])

    # Row 1: Prompt + Model headers
    top_headers = ["Prompt ID", "Prompt"]
    
    # Row 2: Specific column names
    second_headers = ["Prompt ID", "Prompt"]

    # Only add columns for models that actually have data
    models_with_data = []
    for runner_key, label in MODEL_ORDER:
        if runner_key in runner_map and runner_map[runner_key]:
            models_with_data.append((runner_key, label))
            top_headers.extend([label] * len(model_column_headers))
            second_headers.extend(model_column_headers)

    ws.append(top_headers)
    ws.append(second_headers)
    
    # Apply styling to headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for cell in ws[2]:
        cell.font = Font(bold=True)
    
    # Apply model-specific colors to headers
    current_col = 3  # Start after Prompt ID and Prompt columns
    for runner_key, label in models_with_data:
        colors = _get_model_colors(runner_key)
        start_col = current_col
        end_col = start_col + len(model_column_headers) - 1
        
        # Apply header background color to top-level header (row 1)
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")
            cell.font = Font(bold=True, color=colors['header_text'])
        
        # Apply header background color to column headers (row 2)
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")
            cell.font = Font(bold=True, color=colors['header_text'])
        
        current_col = end_col + 1

    # Merge top-level headers
    # Prompt ID and Prompt columns span both rows
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    current_col = 3  # Start after Prompt ID and Prompt columns
    for _, label in models_with_data:
        start_col = current_col
        end_col = start_col + len(model_column_headers) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col).value = label
        current_col = end_col + 1

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    runner_records: Dict[str, List[IterationRecord]] = {}
    max_runs = 0
    for runner_key, records in runner_map.items():
        records_sorted = sorted(
            records,
            key=lambda record: (
                record.iteration if record.iteration is not None else float("inf"),
                record.start_timestamp or "",
                record.end_timestamp or "",
                record.run_id or "",
            ),
        )
        runner_records[runner_key] = records_sorted
        max_runs = max(max_runs, len(records_sorted))

    if max_runs == 0:
        max_runs = 1

    for row_index in range(max_runs):
        row_records: List[Optional[IterationRecord]] = []
        for runner_key, _label in models_with_data:
            records = runner_records.get(runner_key, [])
            row_records.append(records[row_index] if row_index < len(records) else None)

        # Skip this row if all records are None (no data for any model)
        if all(record is None for record in row_records):
            continue

        # Start with Prompt column
        prompt_value = _select_prompt(runner_map) or ""
        prompt_id_value = _select_prompt_id(runner_map) or task_id
        is_primary_row = row_index == 0
        
        # Sanitize prompt for Excel compatibility
        sanitized_prompt = _sanitize_for_excel(prompt_value) if is_primary_row else ""
        
        row = [
            str(prompt_id_value or task_id) if is_primary_row else "",
            sanitized_prompt,
        ]

        hyperlink_targets: List[Tuple[int, str, str]] = []

        for record in row_records:
            if record:
                status = _extract_record_status(record) or ""
                # Use verification_comments from database if available, but only if there's meaningful iteration content
                iteration_uuid = getattr(record, "iteration_uuid", None)
                if not iteration_uuid and isinstance(record.extra, dict):
                    iteration_uuid = (
                        record.extra.get("iteration_uuid")
                        or (record.extra.get("backend_verification_results") or {}).get("iteration_uuid")
                        or (record.extra.get("backend_verification_results") or {}).get("iteration_id")
                    )
                
                # Only show verification comments if there's meaningful iteration content (same logic as iteration UUID)
                if iteration_uuid:
                    auto_comments = getattr(record, 'verification_comments', '') or ""
                else:
                    auto_comments = ""

                model_response = _extract_record_model_response(record) or ""
                if not model_response:
                    model_response = "—"  # Use dash for empty model responses
                
                # Sanitize model response for Excel compatibility
                model_response = _sanitize_for_excel(model_response)

                # Sanitize other text fields for Excel compatibility
                status_label = _sanitize_for_excel(status)
                comments_label = _sanitize_for_excel(auto_comments) if auto_comments else "—"
                execution_uuid = getattr(record, "execution_uuid", None)
                iteration_uuid = getattr(record, "iteration_uuid", None)
                if isinstance(record.extra, dict):
                    execution_uuid = execution_uuid or record.extra.get("execution_uuid")
                    iteration_uuid = (
                        iteration_uuid
                        or record.extra.get("iteration_uuid")
                        or (record.extra.get("backend_verification_results") or {}).get("iteration_uuid")
                        or (record.extra.get("backend_verification_results") or {}).get("iteration_id")
                    )
                hyperlink_display = "—"
                hyperlink_url = None
                if execution_uuid and iteration_uuid:
                    iteration_uuid_str = str(iteration_uuid)
                    hyperlink_url = f"{base_execution_url}/executions/{execution_uuid}/iterations/{iteration_uuid_str}"
                    # Use a more descriptive display text instead of truncated UUID
                    hyperlink_display = f"View Iteration {iteration_uuid_str[:8]}"
            else:
                auto_comments = ""
                model_response = ""
                status_label = ""
                comments_label = ""
                hyperlink_display = "—"
                hyperlink_url = None

            # Format time taken in minutes
            time_taken = ""
            if record and record.duration_seconds is not None:
                # Convert seconds to minutes and format
                minutes = record.duration_seconds / 60.0
                if minutes < 1:
                    time_taken = f"{record.duration_seconds:.1f}s"
                else:
                    time_taken = f"{minutes:.1f}m"
            elif record and record.timelapse:
                time_taken = record.timelapse

            # Extract step count (only if steps data is available for this task)
            steps_value = ""
            if has_steps_data and record and record.total_steps is not None:
                steps_value = str(record.total_steps)

            row_len_before = len(row)
            # Build row based on has_steps_data flag
            row_data = [
                model_response,
                status_label,
                time_taken,
            ]
            if has_steps_data:
                row_data.append(steps_value)
            row_data.extend([
                hyperlink_display,
                auto_comments,  # Comments column (verification comments from database)
            ])
            row.extend(row_data)
            if hyperlink_url:
                # Calculate hyperlink column index based on has_steps_data
                # row_data structure: [model_response, status, time_taken, (steps), hyperlink, comments]
                # Hyperlink is at index 3 (without steps) or 4 (with steps)
                hyperlink_col_offset = 4 if has_steps_data else 3
                hyperlink_targets.append((row_len_before + hyperlink_col_offset + 1, hyperlink_url, hyperlink_display))

        ws.append(row)
        
        # Apply model-specific colors to data rows
        excel_row = ws.max_row
        current_col = 3  # Start after Prompt ID and Prompt columns
        for runner_key, _label in models_with_data:
            colors = _get_model_colors(runner_key)
            start_col = current_col
            end_col = start_col + len(model_column_headers) - 1
            
            # Apply data background color to all cells in this model's columns
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.fill = PatternFill(start_color=colors['data_bg'], end_color=colors['data_bg'], fill_type="solid")
                cell.font = Font(color=colors['data_text'])
            
            current_col = end_col + 1
        
        if hyperlink_targets:
            for col_idx, url, display in hyperlink_targets:
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.value = display
                # Ensure the hyperlink is properly formatted
                if url and url.startswith(('http://', 'https://')):
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
                else:
                    # If URL is malformed, just show the display text without hyperlink
                    cell.value = display
                    cell.font = Font(color="000000")  # Black text, no underline

    # Column widths: Prompt + N models x cols each
    column_widths = [28, 120]  # Prompt ID, Prompt
    # Each model has variable columns based on has_steps_data
    # Set UNIFORM widths for all model columns across all models
    if has_steps_data:
        # Model Response, Pass/Fail, Time Taken, Steps, Iteration Link, Comments
        model_widths = [80, 26, 30, 20, 36, 60]
    else:
        # Model Response, Pass/Fail, Time Taken, Iteration Link, Comments (no Steps)
        model_widths = [80, 26, 30, 36, 60]
    
    for _ in models_with_data:
        column_widths.extend(model_widths)
    
    for idx, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(idx + 1)].width = width

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 160
        for cell in ws[row_idx]:
            cell.alignment = wrap_alignment



# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _safe_load_json(path: Path) -> Optional[Dict[str, object]]:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:  # pragma: no cover - defensive
        logger.error("Unable to parse JSON at %s: %s", path, exc)
        return None


def _aggregate_tool_calls(tool_log_path: Path) -> Tuple[int, Dict[str, int]]:
    if not tool_log_path.exists():
        return 0, {}

    tool_counter: Dict[str, int] = {}
    total = 0
    try:
        with tool_log_path.open(encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    payload = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if payload.get("phase") != "request":
                    continue
                tool_name = payload.get("tool_name") or "unknown"
                tool_counter[tool_name] = tool_counter.get(tool_name, 0) + 1
                total += 1
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("Unable to parse tool log %s: %s", tool_log_path, exc)

    return total, tool_counter


def _extract_iteration_number(name: str) -> Optional[int]:
    match = re.search(r"(iteration[_-]?)(\d+)", name, re.IGNORECASE)
    if match:
        try:
            return int(match.group(2))
        except ValueError:
            return None
    return None


def _safe_float(value: object) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _format_seconds(value: float) -> str:
    if value is None:
        return ""
    seconds = max(0, int(round(value)))
    hours, rem = divmod(seconds, 3600)
    minutes, secs = divmod(rem, 60)
    return f"{hours:02}:{minutes:02}:{secs:02}"


def _sanitize_for_excel(text: str) -> str:
    """Sanitize text for Excel compatibility by removing/replacing problematic characters."""
    if not text:
        return ""
    
    # Replace problematic Unicode characters with ASCII equivalents
    replacements = {
        '•': '-',  # Bullet point to dash
        '✅': '[OK]',  # Checkmark to text
        '❌': '[FAIL]',  # X mark to text
        '⚠️': '[WARN]',  # Warning to text
        '🔍': '[SEARCH]',  # Magnifying glass to text
        '📝': '[NOTE]',  # Memo to text
        '🎯': '[TARGET]',  # Target to text
        '🚀': '[LAUNCH]',  # Rocket to text
        '💡': '[IDEA]',  # Light bulb to text
        '⭐': '[STAR]',  # Star to text
        '🔥': '[HOT]',  # Fire to text
        '💯': '[100]',  # Hundred to text
        '📊': '[CHART]',  # Chart to text
        '📈': '[UP]',  # Trending up to text
        '📉': '[DOWN]',  # Trending down to text
        '🔧': '[TOOL]',  # Wrench to text
        '⚡': '[FAST]',  # Lightning to text
        '🎉': '[CELEBRATE]',  # Party to text
        '🏆': '[TROPHY]',  # Trophy to text
        '📋': '[CLIPBOARD]',  # Clipboard to text
        '🔒': '[LOCK]',  # Lock to text
        '🔓': '[UNLOCK]',  # Unlock to text
        '📌': '[PIN]',  # Pin to text
        '📍': '[LOCATION]',  # Location pin to text
        '⏰': '[TIME]',  # Clock to text
        '📅': '[CALENDAR]',  # Calendar to text
        '📧': '[EMAIL]',  # Email to text
        '📞': '[PHONE]',  # Phone to text
        '🌐': '[WEB]',  # Globe to text
        '💻': '[COMPUTER]',  # Computer to text
        '📱': '[MOBILE]',  # Mobile phone to text
        '🖥️': '[DESKTOP]',  # Desktop computer to text
        '⌨️': '[KEYBOARD]',  # Keyboard to text
        '🖱️': '[MOUSE]',  # Mouse to text
        '💾': '[SAVE]',  # Floppy disk to text
        '📁': '[FOLDER]',  # Folder to text
        '📄': '[DOCUMENT]',  # Document to text
        '📝': '[WRITE]',  # Writing hand to text
        '✏️': '[PENCIL]',  # Pencil to text
        '🖊️': '[PEN]',  # Pen to text
        '🖍️': '[CRAYON]',  # Crayon to text
        '📏': '[RULER]',  # Ruler to text
        '📐': '[TRIANGLE]',  # Triangle ruler to text
        '✂️': '[SCISSORS]',  # Scissors to text
        '📎': '[PAPERCLIP]',  # Paperclip to text
        '📌': '[PUSHPIN]',  # Pushpin to text
        '🔗': '[LINK]',  # Link to text
        '🔖': '[BOOKMARK]',  # Bookmark to text
        '📖': '[BOOK]',  # Open book to text
        '📚': '[BOOKS]',  # Books to text
        '📓': '[NOTEBOOK]',  # Notebook to text
        '📔': '[NOTEBOOK2]',  # Notebook with decorative cover to text
        '📒': '[LEDGER]',  # Ledger to text
        '📕': '[RED_BOOK]',  # Red book to text
        '📗': '[GREEN_BOOK]',  # Green book to text
        '📘': '[BLUE_BOOK]',  # Blue book to text
        '📙': '[ORANGE_BOOK]',  # Orange book to text
        '📃': '[PAGE]',  # Page with curl to text
        '📄': '[PAGE_FACING_UP]',  # Page facing up to text
        '📰': '[NEWSPAPER]',  # Newspaper to text
        '🗞️': '[ROLLED_NEWSPAPER]',  # Rolled-up newspaper to text
        '📑': '[BOOKMARK_TABS]',  # Bookmark tabs to text
        '🔖': '[BOOKMARK2]',  # Bookmark to text
        '🏷️': '[LABEL]',  # Label to text
        '💰': '[MONEY]',  # Money bag to text
        '💳': '[CREDIT_CARD]',  # Credit card to text
        '💎': '[DIAMOND]',  # Diamond to text
        '⚖️': '[BALANCE]',  # Balance scale to text
        '🔧': '[WRENCH]',  # Wrench to text
        '🔨': '[HAMMER]',  # Hammer to text
        '⚒️': '[HAMMER_PICK]',  # Hammer and pick to text
        '🛠️': '[HAMMER_WRENCH]',  # Hammer and wrench to text
        '⛏️': '[PICK]',  # Pick to text
        '🔩': '[NUT_BOLT]',  # Nut and bolt to text
        '⚙️': '[GEAR]',  # Gear to text
        '🗜️': '[CLAMP]',  # Clamp to text
        '⚗️': '[ALEMBIC]',  # Alembic to text
        '🔬': '[MICROSCOPE]',  # Microscope to text
        '🔭': '[TELESCOPE]',  # Telescope to text
        '📡': '[SATELLITE]',  # Satellite antenna to text
        '💉': '[SYRINGE]',  # Syringe to text
        '💊': '[PILL]',  # Pill to text
        '🩹': '[BANDAGE]',  # Adhesive bandage to text
        '🩺': '[STETHOSCOPE]',  # Stethoscope to text
        '🚪': '[DOOR]',  # Door to text
        '🛏️': '[BED]',  # Bed to text
        '🛋️': '[COUCH]',  # Couch and lamp to text
        '🚽': '[TOILET]',  # Toilet to text
        '🚿': '[SHOWER]',  # Shower to text
        '🛁': '[BATHTUB]',  # Bathtub to text
        '🧴': '[LOTION]',  # Lotion bottle to text
        '🧷': '[SAFETY_PIN]',  # Safety pin to text
        '🧹': '[BROOM]',  # Broom to text
        '🧺': '[BASKET]',  # Basket to text
        '🧻': '[ROLL_PAPER]',  # Roll of paper to text
        '🧼': '[SOAP]',  # Soap to text
        '🧽': '[SPONGE]',  # Sponge to text
        '🧯': '[FIRE_EXTINGUISHER]',  # Fire extinguisher to text
        '🧲': '[MAGNET]',  # Magnet to text
        '🧳': '[LUGGAGE]',  # Luggage to text
        '🧴': '[LOTION2]',  # Lotion bottle to text
        '🧵': '[THREAD]',  # Thread to text
        '🧶': '[YARN]',  # Yarn to text
        '🧷': '[SAFETY_PIN2]',  # Safety pin to text
        '🧸': '[TEDDY_BEAR]',  # Teddy bear to text
        '🧹': '[BROOM2]',  # Broom to text
        '🧺': '[BASKET2]',  # Basket to text
        '🧻': '[ROLL_PAPER2]',  # Roll of paper to text
        '🧼': '[SOAP2]',  # Soap to text
        '🧽': '[SPONGE2]',  # Sponge to text
        '🧯': '[FIRE_EXTINGUISHER2]',  # Fire extinguisher to text
        '🧲': '[MAGNET2]',  # Magnet to text
        '🧳': '[LUGGAGE2]',  # Luggage to text
    }
    
    # Apply replacements
    sanitized = text
    for unicode_char, ascii_replacement in replacements.items():
        sanitized = sanitized.replace(unicode_char, ascii_replacement)
    
    # Remove any remaining control characters (except newlines and tabs)
    sanitized = ''.join(char for char in sanitized if ord(char) >= 32 or char in '\n\t')
    
    # Truncate if too long (Excel cell limit is 32,767 characters)
    if len(sanitized) > 32000:
        sanitized = sanitized[:32000] + "... [TRUNCATED]"
    
    return sanitized


def _safe_sheet_name(workbook: Workbook, task_id: str) -> str:
    base = task_id[:28] if len(task_id) > 28 else task_id
    candidate = base or "Task"
    suffix = 1
    existing = {ws.title for ws in workbook.worksheets}
    while candidate in existing:
        candidate = f"{base[:25]}_{suffix}" if len(base) > 25 else f"{base}_{suffix}"
        suffix += 1
    return candidate


def _safe_parse_timestamp(value: str) -> Optional[datetime]:
    try:
        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S,%f")
    except Exception:
        return None


def _select_prompt(runner_map: Dict[str, List[IterationRecord]]) -> Optional[str]:
    for runner_key, _ in MODEL_ORDER:
        for record in runner_map.get(runner_key, []):
            if record.prompt:
                return record.prompt
    for records in runner_map.values():
        for record in records:
            if record.prompt:
                return record.prompt
    return None


def _select_prompt_id(runner_map: Dict[str, List[IterationRecord]]) -> Optional[str]:
    for runner_key, _ in MODEL_ORDER:
        for record in runner_map.get(runner_key, []):
            if record.prompt_id:
                return str(record.prompt_id)
    for records in runner_map.values():
        for record in records:
            if record.prompt_id:
                return str(record.prompt_id)
    return None


def _estimate_expected_tool_calls(runner_map: Dict[str, List[IterationRecord]]) -> str:
    values: List[int] = []
    for records in runner_map.values():
        for record in records:
            if record.tool_calls_total is not None:
                values.append(int(record.tool_calls_total))
            elif record.extra:
                usage = record.extra.get("tool_usage") or {}
                total = usage.get("total")
                if total:
                    try:
                        values.append(int(total))
                    except Exception:
                        continue
    if values:
        return str(max(values))
    return ""


def _aggregate_runner_stats(records: List[IterationRecord]) -> Dict[str, object]:
    stats = {
        "total": 0,
        "pass_count": 0,
        "fail_count": 0,
        "crash_count": 0,
        "timeout_count": 0,
        "other_count": 0,
        "tool_calls_total": 0,
        "tool_calls_by_tool": Counter(),
        "completion_reasons": [],
        "status_reasons": [],
        "errors": [],
        "records": records,
    }

    normalized_records: List[Tuple[IterationRecord, str]] = []
    non_crash_present = False

    for record in records:
        normalized = _infer_status(
            record.status,
            extra=record.extra,
            status_reason=record.status_reason,
            completion_reason=record.completion_reason,
        )
        normalized = (normalized or "").upper()
        normalized_records.append((record, normalized))
        if normalized and normalized != "CRASHED":
            non_crash_present = True

    records_for_stats: List[Tuple[IterationRecord, str]] = []
    for record, normalized in normalized_records:
        # Only include PASSED and FAILED iterations in reports
        if normalized in ["PASSED", "FAILED"]:
            records_for_stats.append((record, normalized))
        # Exclude all other statuses: CRASHED, PENDING, EXECUTING, UNKNOWN, TIMEOUT, etc.
        # (TIMEOUT is now converted to FAILED in _simplify_status, so it won't reach here)

    stats["records"] = [record for record, _ in records_for_stats]

    for record, normalized_status in records_for_stats:
        status = normalized_status.lower()
        stats["total"] += 1
        if status == "passed":
            stats["pass_count"] += 1
        elif status == "failed":
            stats["fail_count"] += 1
        # All other statuses are excluded from reports, so we don't count them
        # (crash_count, timeout_count, other_count remain 0)

        if record.tool_calls_total is not None:
            stats["tool_calls_total"] += record.tool_calls_total

        if record.tool_calls_by_tool:
            stats["tool_calls_by_tool"].update(
                {tool: int(count) for tool, count in record.tool_calls_by_tool.items()}
            )
        elif record.extra:
            usage = record.extra.get("tool_usage") or {}
            by_tool = usage.get("by_tool") or {}
            stats["tool_calls_by_tool"].update({tool: int(count) for tool, count in by_tool.items()})
            if record.tool_calls_total is None:
                stats["tool_calls_total"] += sum(int(count) for count in by_tool.values())

        if record.completion_reason:
            stats["completion_reasons"].append(record.completion_reason)
        if record.status_reason:
            stats["status_reasons"].append(record.status_reason)
        if record.extra:
            if record.extra.get("error"):
                stats["errors"].append(str(record.extra["error"]))
            if record.extra.get("status_reason"):
                stats["status_reasons"].append(str(record.extra["status_reason"]))

    return stats


def _format_breaking_string(stats: Dict[str, object]) -> str:
    total = stats.get("total", 0)
    if not total:
        return ""  # Return empty string instead of "No data" for crashed/empty iterations
    break_count = total - stats.get("pass_count", 0)
    if break_count:
        return f"Yes, {break_count}/{total}"
    return f"No, 0/{total}"


def _difficulty_from_runner_stats(runner_map: Dict[str, List[IterationRecord]]) -> str:
    """
    Calculate difficulty based on Pass@8 criteria using all iterations across all models.
    
    Difficulty Criteria:
    - Easy: Pass@8 = 100% (0 failures out of 8 iterations)
    - Medium: 40% < Pass@8 < 100% (1-4 failures out of 8 iterations)  
    - Hard: Pass@8 < 40% (5-8 failures out of 8 iterations)
    """
    # Collect all iterations across all runners for this task
    all_records: List[IterationRecord] = []
    for runner_records in runner_map.values():
        all_records.extend(runner_records)
    
    if not all_records:
        return "Unknown"
    
    # Count only PASSED and FAILED iterations (exclude CRASHED, PENDING, EXECUTING, etc.)
    passed_count = 0
    failed_count = 0
    
    for record in all_records:
        normalized = _infer_status(
            record.status,
            extra=record.extra,
            status_reason=record.status_reason,
            completion_reason=record.completion_reason,
        )
        normalized = (normalized or "").upper()
        
        if normalized == "PASSED":
            passed_count += 1
        elif normalized == "FAILED":
            failed_count += 1
        # Exclude all other statuses (CRASHED, PENDING, EXECUTING, etc.)
    
    total_iterations = passed_count + failed_count
    
    if total_iterations == 0:
        return "Unknown"
    
    # Calculate Pass@8 percentage
    pass_percentage = (passed_count / total_iterations) * 100
    
    # Apply Pass@8 criteria
    if pass_percentage == 100.0:
        return "Easy"  # Pass@8 = 100% (0 failures)
    elif pass_percentage > 40.0:
        return "Medium"  # 40% < Pass@8 < 100% (1-4 failures out of 8)
    else:
        return "Hard"  # Pass@8 < 40% (5-8 failures out of 8)


def _summarize_runner(runner_key: str, records: List[IterationRecord]) -> Dict[str, object]:
    label = dict(MODEL_ORDER).get(
        runner_key, KNOWN_RUNNERS.get(runner_key, runner_key.title())
    )
    records_sorted = sorted(records, key=lambda r: r.iteration)
    stats = _aggregate_runner_stats(records_sorted)

    # Simplified format - just the content without "Iteration X" prefix
    model_response = _format_iteration_details_simple(
        records_sorted, _extract_record_model_response
    )
    # Simplified status format - just the status values without iteration labels
    pass_fail_text = ", ".join([_extract_record_status(r) or "" for r in records_sorted]) if records_sorted else ""

    return {
        "runner_key": runner_key,
        "runner_label": label,
        "total_iterations": stats["total"],
        "pass_count": stats["pass_count"],
        "fail_count": stats["fail_count"],
        "crash_count": stats["crash_count"],
        "timeout_count": stats["timeout_count"],
        "tool_calls_total": stats["tool_calls_total"],
        "model_response": model_response,
        "pass_fail_text": pass_fail_text,
        "comments": "",
    }


def _format_iteration_details(
    records: List[IterationRecord],
    value_getter,
    *,
    include_status: bool = True,
) -> str:
    if not records:
        return ""

    parts: List[str] = []
    for record in records:
        status = (record.status or "unknown").upper()
        label = f"Iteration {record.iteration}"
        if include_status:
            label += f" ({status})"
        value = value_getter(record)
        value_str = str(value).strip() if value else ""
        parts.append(f"{label}:\n{value_str}")

    return "\n\n".join(parts)


def _format_iteration_details_simple(
    records: List[IterationRecord],
    value_getter,
) -> str:
    """Format iteration details without 'Iteration X' prefix - just the content"""
    if not records:
        return ""

    parts: List[str] = []
    for record in records:
        value = value_getter(record)
        value_str = str(value).strip() if value else ""
        if value_str:
            parts.append(value_str)

    return "\n\n".join(parts) if parts else ""


def _normalize_status_token(status: Optional[str]) -> Optional[str]:
    if not status:
        return None

    token = status.strip().lower()
    if not token:
        return None

    mapping = {
        "passed": "PASSED",
        "pass": "PASSED",
        "success": "PASSED",
        "succeeded": "PASSED",
        "completed": "PASSED",
        "complete": "PASSED",
        "done": "PASSED",
        "ok": "PASSED",
        "verified": "PASSED",
        "fail": "FAILED",
        "failed": "FAILED",
        "failure": "FAILED",
        "verification_failed": "FAILED",
        "verification-failed": "FAILED",
        "verification error": "FAILED",
        "timeout": "TIMEOUT",
        "timed_out": "TIMEOUT",
        "timed-out": "TIMEOUT",
        "timed out": "TIMEOUT",
        "time_out": "TIMEOUT",
        "crash": "CRASHED",
        "crashed": "CRASHED",
        "error": "CRASHED",
        "errored": "CRASHED",
        "exception": "CRASHED",
        "unknown": "UNKNOWN",
        "pending": "PENDING",
        "in_progress": "EXECUTING",
        "executing": "EXECUTING",
        "running": "EXECUTING",
    }

    if token in mapping:
        return mapping[token]

    # Preserve uppercase version for any other explicit status values
    return token.upper()


def _infer_status(
    status: Optional[str],
    *,
    extra: Optional[Dict[str, object]] = None,
    status_reason: Optional[str] = None,
    completion_reason: Optional[str] = None,
) -> str:
    extra = extra or {}

    candidates = [status, extra.get("status"), extra.get("verification_status"), extra.get("state"), extra.get("execution_status")]

    for candidate in candidates:
        normalized = _normalize_status_token(candidate)
        if normalized and normalized != "UNKNOWN":
            # Convert to simplified status classification
            return _simplify_status(normalized, status_reason, completion_reason, extra)

    # Don't do text analysis - trust the database status
    # If we still could not infer, fall back to normalized token
    normalized = _normalize_status_token(status)
    if normalized:
        return _simplify_status(normalized, status_reason, completion_reason, extra)
    return "UNKNOWN"


def _simplify_status(
    status: str,
    status_reason: Optional[str] = None,
    completion_reason: Optional[str] = None,
    extra: Optional[Dict[str, object]] = None,
) -> str:
    """
    Simplify status classification to only PASSED, FAILED, or exclude categories.
    
    Rules:
    - PASSED, SUCCESS, COMPLETED, etc. -> PASSED
    - FAILED, VERIFICATION_FAILED, etc. -> FAILED  
    - TIMEOUT, TIMED_OUT, etc. -> FAILED (timeout is treated as failed)
    - CRASHED -> CRASHED (will be excluded unless it actually passed/failed)
    - PENDING -> PENDING (will be excluded)
    - EXECUTING -> EXECUTING (will be excluded)
    - UNKNOWN -> UNKNOWN (will be excluded)
    """
    status_upper = status.upper()
    
    # Map all pass-related statuses to PASSED
    if status_upper in ["PASSED", "SUCCESS", "SUCCEEDED", "COMPLETED", "COMPLETE", "DONE", "OK", "VERIFIED"]:
        return "PASSED"
    
    # Map all fail-related statuses to FAILED
    elif status_upper in ["FAILED", "FAIL", "FAILURE", "VERIFICATION_FAILED", "VERIFICATION-FAILED", "VERIFICATION ERROR"]:
        return "FAILED"
    
    # Map all timeout-related statuses to FAILED
    elif status_upper in ["TIMEOUT", "TIMED_OUT", "TIMED-OUT", "TIMED OUT", "TIME_OUT"]:
        return "FAILED"  # Timeout is treated as failed
    
    # For these statuses, return as-is (they will be excluded from reports)
    elif status_upper in ["CRASHED", "PENDING", "EXECUTING", "UNKNOWN"]:
        # Don't do text analysis - trust the database status
        # These statuses will be excluded from reports
        return status_upper
    
    # For any other status, return as-is
    return status_upper


def _extract_task_completion_response(model_response: str) -> str:
    """Extract the natural response from the model.
    
    Args:
        model_response: The full model response text
        
    Returns:
        The natural response from the model
    """
    if not model_response or not isinstance(model_response, str):
        return model_response or ""
    
    # Return the full natural response without any special parsing
    return model_response


def _clean_task_completion_start(response: str) -> str:
    """Clean only the beginning of the task completion response.
    
    Removes leading markdown formatting and excessive whitespace,
    but preserves the actual response content.
    """
    if not response:
        return response
    
    import re
    
    # Remove leading markdown formatting like **, ##, etc. and whitespace
    response = re.sub(r'^\s*\*{1,2}\s*', '', response)
    response = re.sub(r'^\s*#+\s*', '', response)
    
    # Remove leading dashes, underscores, or other common separators
    response = re.sub(r'^[\s\-_=]+\s*', '', response)
    
    # Remove leading whitespace and newlines
    response = response.lstrip()
    
    return response


def _extract_record_model_response(record: IterationRecord) -> str:
    """Extract model response with priority order.
    
    Priority order:
    1. Check database field (if already stored)
    2. First priority: Conversation history files
    3. Second priority: task_responses files (last item's "message")
    4. Third priority: action_timeline.json
    
    Note: This function only READS data. DB writes should happen during task execution.
    """
    
    # Initially check database - if available, return immediately
    if hasattr(record, 'last_model_response') and record.last_model_response:
        return record.last_model_response
    
    # Fallback to file-based extraction for legacy records
    if not record.iteration_directory:
        # Fallback to standard fields if no iteration directory
        if record.completion_reason:
            return record.completion_reason
        if record.status_reason:
            return record.status_reason
        extra = record.extra or {}
        if extra.get("status_reason"):
            return str(extra["status_reason"])
        if extra.get("error"):
            return str(extra["error"])
        return "No response captured."
    
    runner_path = Path(record.iteration_directory)
    found_response = None
    
    # First priority: Look in conversation_history for files ending with _task_execution_conversation.json
    conversation_dir = runner_path / "conversation_history"
    if conversation_dir.exists():
        for conv_file in sorted(conversation_dir.glob("*_task_execution_conversation.json")):
            conv_data = _safe_load_json(conv_file) or {}
            # Handle both direct message arrays and conversation_flow format
            messages = conv_data if isinstance(conv_data, list) else conv_data.get("messages", [])
            conversation_flow = conv_data.get("conversation_flow", [])
            
            # Try conversation_flow first (newer format)
            if conversation_flow:
                for item in reversed(conversation_flow):
                    if item.get("role") == "assistant" and item.get("content"):
                        content = item["content"]
                        if content and content.strip():
                            found_response = _extract_task_completion_response(content)  # Full message, no truncation
                            break
            
            # Fallback to messages array
            if not found_response:
                for msg in reversed(messages):
                    if msg.get("role") == "assistant" and msg.get("content"):
                        content = msg["content"]
                        if isinstance(content, str) and content.strip():
                            found_response = _extract_task_completion_response(content)  # Full message, no truncation
                            break
                        elif isinstance(content, list):
                            for block in content:
                                if isinstance(block, dict) and block.get("type") == "text":
                                    text = block.get("text", "")
                                    if text and text.strip():
                                        found_response = _extract_task_completion_response(text)  # Full message
                                        break
                            if found_response:
                                break
            break
    
    # Second priority: Check task_responses for "message" in last item
    if not found_response:
        task_responses_dir = runner_path / "task_responses"
        if task_responses_dir.exists():
            response_files = sorted(task_responses_dir.glob("*.json"))
            if response_files:
                # Get the last (most recent) file
                last_response_file = response_files[-1]
                try:
                    response_data = _safe_load_json(last_response_file) or {}
                    # Check if there's an item_summary array
                    item_summary = response_data.get("item_summary", [])
                    if item_summary and isinstance(item_summary, list):
                        # Get the last item
                        last_item = item_summary[-1]
                        if isinstance(last_item, dict):
                            # Look for "message" field
                            message = last_item.get("message") or last_item.get("content")
                            if message and isinstance(message, str) and message.strip():
                                found_response = message.strip()
                except Exception as e:
                    logger.warning(f"Failed to extract from task_responses: {e}")
    
    # Third priority: Check action_timeline.json
    if not found_response:
        action_timeline_file = runner_path / "action_timeline.json"
        if action_timeline_file.exists():
            try:
                timeline_data = _safe_load_json(action_timeline_file) or {}
                entries = timeline_data.get("entries", [])
                # Look for model_response entries (go in reverse to get the last one)
                for entry in reversed(entries):
                    if entry.get("entry_type") == "model_response" and entry.get("content"):
                        content = entry["content"]
                        if content and content.strip():
                            found_response = content.strip()
                            break
            except Exception as e:
                logger.warning(f"Failed to extract from action_timeline.json: {e}")
    
    # Return the found response (no DB writes in report generation)
    if found_response:
        return found_response
    
    # Fallback to standard fields if nothing found in files
    if record.completion_reason:
        return record.completion_reason
    if record.status_reason:
        return record.status_reason
    extra = record.extra or {}
    if extra.get("status_reason"):
        return str(extra["status_reason"])
    if extra.get("error"):
        return str(extra["error"])
    
    return "No response captured."


def _get_record_tool_usage_map(record: IterationRecord) -> Dict[str, int]:
    if record.tool_calls_by_tool:
        return {tool: int(count) for tool, count in record.tool_calls_by_tool.items()}
    usage = (record.extra or {}).get("tool_usage") or {}
    by_tool = usage.get("by_tool") or {}
    return {tool: int(count) for tool, count in by_tool.items()}


def _extract_record_tool_usage(record: IterationRecord) -> str:
    usage_map = _get_record_tool_usage_map(record)
    if not usage_map:
        return "No data"
    return "; ".join(f"{tool} x{count}" for tool, count in usage_map.items())


def _extract_record_tool_calls(record: IterationRecord) -> str:
    total = record.tool_calls_total
    if total is None:
        usage_map = _get_record_tool_usage_map(record)
        if usage_map:
            total = sum(usage_map.values())
    if total is None:
        return "No data"
    return str(total)


def _extract_record_status(record: IterationRecord) -> str:
    normalized = _infer_status(
        record.status,
        extra=record.extra,
        status_reason=record.status_reason,
        completion_reason=record.completion_reason,
    )
    return normalized if normalized != "UNKNOWN" else ""




def _format_tool_usage(counter: Counter) -> str:
    if not counter:
        return "None"
    parts = [f"{tool} x{count}" for tool, count in counter.most_common()]
    return "; ".join(parts)


def _dedupe_strings(values: Iterable[str], limit: int = 5) -> List[str]:
    seen = set()
    result = []
    for value in values:
        if not value:
            continue
        trimmed = value.strip()
        if not trimmed:
            continue
        if trimmed in seen:
            continue
        seen.add(trimmed)
        result.append(trimmed)
        if len(result) >= limit:
            break
    return result


def _build_snapshot(
    summary_rows: List[Dict[str, object]],
    iterations: List[IterationRecord],
    task_rows: Dict[str, Dict[str, List[IterationRecord]]],
) -> Dict[str, object]:
    serialized_iterations = [_serialize_iteration(record) for record in iterations]
    tasks = sorted(task_rows.keys())
    # Only include models that actually have data in this batch
    models_with_data = set()
    for task_id, runner_map in task_rows.items():
        for runner_key in runner_map.keys():
            if runner_key in dict(MODEL_ORDER):
                models_with_data.add(runner_key)
    
    models_with_data_labels = [label for runner_key, label in MODEL_ORDER if runner_key in models_with_data]
    
    filters = {
        "models": models_with_data_labels,
        "runners": models_with_data_labels,
        "tasks": tasks,
    }

    summary_table: List[Dict[str, object]] = []
    single_task_tables: Dict[str, List[Dict[str, object]]] = {}

    for summary_row in summary_rows:
        summary_table.append({
            "Prompt ID": summary_row.get("Prompt ID"),
            "Task": summary_row.get("Task"),
            "Prompt": summary_row.get("Prompt"),
            "Difficulty": summary_row.get("Difficulty"),
            "TotalTimeSeconds": summary_row.get("Total Time Seconds"),
            "TotalTimeFormatted": _format_seconds(summary_row.get("Total Time Seconds")) if summary_row.get("Total Time Seconds") is not None else None,
            "TaskStartTime": summary_row.get("Task Start Time"),
            "TaskEndTime": summary_row.get("Task End Time"),
            "AverageIterationMinutes": summary_row.get("Average Iteration Time Minutes"),
        })

    for task_id, runner_map in task_rows.items():
        task_iterations: List[Dict[str, object]] = []
        for runner_key, records in runner_map.items():
            for record in records:
                task_iterations.append({
                    "Task": task_id,
                    "Iteration": record.iteration,
                    "Prompt": record.prompt,
                    "PromptId": record.prompt_id,
                    "RunnerKey": runner_key,
                    "RunnerLabel": KNOWN_RUNNERS.get(runner_key, runner_key.title()),
                    "Status": record.status,
                    "ExecutionTimeSeconds": record.duration_seconds,
                    "ExecutionTimeFormatted": _format_seconds(record.duration_seconds) if record.duration_seconds is not None else None,
                    "StartTime": record.start_timestamp,
                    "EndTime": record.end_timestamp,
                })

        task_iterations.sort(key=lambda row: (row.get("Iteration", 0), row.get("RunnerLabel", "")))
        single_task_tables[task_id] = task_iterations

    tasks_map = {}
    for task_id, runner_map in task_rows.items():
        runner_summary = {}
        # Only include models that actually have data for this task
        for runner_key, label in MODEL_ORDER:
            if runner_key in runner_map and runner_map[runner_key]:
                runner_summary[label] = _summarize_runner(
                    runner_key, runner_map[runner_key]
                )
        tasks_map[task_id] = runner_summary

    return {
        "summary": summary_rows,
        "summary_table": summary_table,
        "iterations": serialized_iterations,
        "tasks": tasks_map,
        "single_task_tables": single_task_tables,
        "filters": filters,
    }


def _serialize_iteration(record: IterationRecord) -> Dict[str, object]:
    return {
        "task_id": record.task_id,
        "runner": record.runner_label,
        "runner_key": record.runner,
        "iteration": record.iteration,
        "status": record.status,
        "status_reason": record.status_reason,
        "completion_reason": record.completion_reason,
        "duration_seconds": record.duration_seconds,
        "timelapse": record.timelapse,
        "file_timelapse_seconds": record.file_timelapse_seconds,
        "tool_calls_total": record.tool_calls_total,
        "tool_calls_by_tool": record.tool_calls_by_tool,
        "unique_tools": record.unique_tools,
        "prompt": record.prompt,
        "prompt_id": record.prompt_id,
        "model": record.model,
        "run_id": record.run_id,
        "start_timestamp": record.start_timestamp,
        "end_timestamp": record.end_timestamp,
        "iteration_directory": record.iteration_directory,
        "extra": record.extra,
    }


def _populate_insights_sheet(ws, iterations: List[IterationRecord], task_rows: Dict[str, Dict[str, List[IterationRecord]]], batch_insights: Dict[str, str] = None) -> None:
    """Populate the Insights sheet with batch-level and execution-level insights"""
    
    # Get unique models from the data
    models_in_data = set()
    for record in iterations:
        if record.runner:
            models_in_data.add(record.runner)
    
    # Sort models according to MODEL_ORDER
    models_sorted = []
    for model_key, model_label in MODEL_ORDER:
        if model_key in models_in_data:
            models_sorted.append((model_key, model_label))
    
    # Build headers: Task Identifier + Model columns
    headers = ["Task Identifier"]
    headers.extend([model_label for _, model_label in models_sorted])
    
    ws.append(headers)
    
    # Apply header styling
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add Overview row with batch-level insights
    overview_row = ["Overview"]
    for model_key, model_label in models_sorted:
        if batch_insights and model_key in batch_insights:
            insights = batch_insights[model_key] or "No batch insights available"
            overview_row.append(insights)
        else:
            overview_row.append("No batch insights available")
    ws.append(overview_row)
    
    # Add execution-level insights for each task
    for task_id, runner_map in sorted(task_rows.items()):
        task_row = [task_id]
        
        for model_key, model_label in models_sorted:
            if model_key in runner_map and runner_map[model_key]:
                # Get execution-level insights (from the first iteration or combine all)
                first_record = runner_map[model_key][0]
                insights = first_record.eval_insights or "No insights available"
                task_row.append(insights)
            else:
                task_row.append("No data")
        
        ws.append(task_row)
    
    # Auto-adjust column widths and row heights
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)  # Cap at 80 characters for better readability
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set row heights for better readability (similar to model responses)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row_idx in range(2, ws.max_row + 1):  # Skip header row
        ws.row_dimensions[row_idx].height = 80  # Increased height for insights
        for cell in ws[row_idx]:
            cell.alignment = wrap_alignment


def _write_json_snapshot(path: Path, payload: Dict[str, object]) -> None:
    try:
        with path.open("w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, default=str)
        logger.info("🧾 JSON snapshot saved: %s", path)
    except Exception as exc:
        logger.warning("⚠️ Unable to write JSON snapshot %s: %s", path, exc)
