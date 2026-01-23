"""
Tests for Excel-based access control module.

Tests the AccessControl class which loads admin/user roles from an Excel file.
"""
import tempfile
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook

from app.core.access_control import AccessControl


def _reset_access_control_cache():
    """Reset internal cache between tests."""
    AccessControl._roles_by_email = {}
    AccessControl._last_mtime = None
    AccessControl._enabled = False


def _write_access_sheet(path: Path, admins: list[str] = None, users: list[str] = None):
    """Helper to create a test Excel file with Admins and Users sheets."""
    wb = Workbook()
    
    # Admins sheet
    admins_ws = wb.active
    admins_ws.title = "Admins"
    admins_ws.append(["name", "email"])
    for email in (admins or []):
        admins_ws.append(["Admin User", email])
    
    # Users sheet
    users_ws = wb.create_sheet("Users")
    users_ws.append(["name", "email"])
    for email in (users or []):
        users_ws.append(["Normal User", email])
    
    wb.save(path)


@pytest.mark.unit
class TestAccessControlDisabled:
    """Test access control behavior when Excel file is missing."""
    
    def test_disabled_when_file_missing(self, monkeypatch):
        """If the Excel file does not exist, access control is disabled."""
        _reset_access_control_cache()
        missing_path = Path(tempfile.gettempdir()) / "nonexistent_access_sheet.xlsx"
        monkeypatch.setattr("app.core.access_control.settings.ACCESS_CONTROL_EXCEL_PATH", str(missing_path))
        
        assert AccessControl.is_enabled() is False
        assert AccessControl.get_role_for_email("admin@example.com") is None
        assert AccessControl.is_admin_email("admin@example.com") is False
        assert AccessControl.is_user_email("user@example.com") is False


@pytest.mark.unit
class TestAccessControlRoleResolution:
    """Test role resolution from Excel file."""
    
    def test_loads_admin_and_user_roles(self, monkeypatch):
        """Admins and Users sheets map to 'admin' and 'user' roles, case-insensitive."""
        _reset_access_control_cache()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        
        try:
            _write_access_sheet(
                path,
                admins=["admin@example.com", "ADMIN2@EXAMPLE.COM"],
                users=["user@example.com"]
            )
            monkeypatch.setattr("app.core.access_control.settings.ACCESS_CONTROL_EXCEL_PATH", str(path))
            
            # First call should load from disk
            assert AccessControl.is_enabled() is True
            
            # Admin emails (case-insensitive)
            assert AccessControl.get_role_for_email("admin@example.com") == "admin"
            assert AccessControl.get_role_for_email("ADMIN@EXAMPLE.COM") == "admin"
            assert AccessControl.get_role_for_email("Admin@Example.Com") == "admin"
            assert AccessControl.get_role_for_email("ADMIN2@example.com") == "admin"
            
            # User email
            assert AccessControl.get_role_for_email("user@example.com") == "user"
            assert AccessControl.get_role_for_email("USER@EXAMPLE.COM") == "user"
            
            # Convenience helpers
            assert AccessControl.is_admin_email("admin@example.com") is True
            assert AccessControl.is_admin_email("ADMIN2@example.com") is True
            assert AccessControl.is_user_email("user@example.com") is True
            
            # Unknown email
            assert AccessControl.get_role_for_email("other@example.com") is None
            assert AccessControl.is_admin_email("other@example.com") is False
            assert AccessControl.is_user_email("other@example.com") is False
        finally:
            if path.exists():
                path.unlink()
    
    def test_uses_cache_until_file_changes(self, monkeypatch):
        """Subsequent calls reuse cache until the Excel file mtime changes."""
        _reset_access_control_cache()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        
        try:
            # Initial version: only admin@example.com
            _write_access_sheet(path, admins=["admin@example.com"], users=[])
            monkeypatch.setattr("app.core.access_control.settings.ACCESS_CONTROL_EXCEL_PATH", str(path))
            
            # Load first version
            assert AccessControl.get_role_for_email("admin@example.com") == "admin"
            assert AccessControl.get_role_for_email("user@example.com") is None
            
            # Overwrite file to add user@example.com
            _write_access_sheet(
                path,
                admins=["admin@example.com"],
                users=["user@example.com"]
            )
            
            # Force mtime change
            path.touch()
            
            # Now cache should refresh and see user@example.com as 'user'
            assert AccessControl.get_role_for_email("user@example.com") == "user"
        finally:
            if path.exists():
                path.unlink()
    
    def test_handles_missing_sheets_gracefully(self, monkeypatch):
        """If Admins or Users sheet is missing, only existing sheets are loaded."""
        _reset_access_control_cache()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        
        try:
            # Only Admins sheet, no Users sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Admins"
            ws.append(["name", "email"])
            ws.append(["Admin", "admin@example.com"])
            wb.save(path)
            
            monkeypatch.setattr("app.core.access_control.settings.ACCESS_CONTROL_EXCEL_PATH", str(path))
            
            assert AccessControl.is_enabled() is True
            assert AccessControl.get_role_for_email("admin@example.com") == "admin"
            assert AccessControl.get_role_for_email("user@example.com") is None
        finally:
            if path.exists():
                path.unlink()
    
    def test_handles_empty_email_column(self, monkeypatch):
        """Empty email cells are skipped."""
        _reset_access_control_cache()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Admins"
            ws.append(["name", "email"])
            ws.append(["Admin 1", "admin@example.com"])
            ws.append(["Admin 2", ""])  # Empty email
            ws.append(["Admin 3", None])  # None email
            ws.append(["Admin 4", "admin2@example.com"])
            wb.save(path)
            
            monkeypatch.setattr("app.core.access_control.settings.ACCESS_CONTROL_EXCEL_PATH", str(path))
            
            assert AccessControl.is_enabled() is True
            assert AccessControl.get_role_for_email("admin@example.com") == "admin"
            assert AccessControl.get_role_for_email("admin2@example.com") == "admin"
            # Empty emails should not be in the map
            assert len([k for k in AccessControl._roles_by_email.keys() if "admin" in k]) == 2
        finally:
            if path.exists():
                path.unlink()


@pytest.mark.unit
class TestAccessControlRealFile:
    """Test with the actual Access_sheet.xlsx file if it exists."""
    
    def test_loads_real_excel_file(self):
        """Test loading the actual Access_sheet.xlsx file."""
        _reset_access_control_cache()
        
        # Try to load the real file
        real_path = Path(__file__).parent.parent.parent / "access" / "Access_sheet.xlsx"
        
        if real_path.exists():
            # Temporarily override the path
            with patch("app.core.access_control.settings.ACCESS_CONTROL_EXCEL_PATH", str(real_path)):
                # Force reload
                AccessControl._load_roles()
                
                assert AccessControl.is_enabled() is True
                print(f"\n✅ Real Excel file loaded successfully")
                print(f"   Total emails: {len(AccessControl._roles_by_email)}")
                print(f"   Admin emails: {sum(1 for r in AccessControl._roles_by_email.values() if r == 'admin')}")
                print(f"   User emails: {sum(1 for r in AccessControl._roles_by_email.values() if r == 'user')}")
                
                # Test a few lookups
                for email, role in list(AccessControl._roles_by_email.items())[:5]:
                    print(f"   {email} -> {role}")
        else:
            pytest.skip(f"Real Excel file not found at {real_path}")

